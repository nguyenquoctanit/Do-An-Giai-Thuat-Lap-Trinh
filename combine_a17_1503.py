import socket
from time import sleep

import glob
import os

import cv2
import threading
import torch
import numpy as np 
import time

import PySimpleGUI as sg

from PIL import Image,ImageTk
import os
import datetime 
import shutil

from PIL import Image

import traceback

import sqlite3

import multiprocessing

import openpyxl
import datetime

from openpyxl.styles import Alignment
from openpyxl import Workbook
from datetime import date
from openpyxl.styles import Font

def socket_connect(host, port):
    try:
        soc.connect((host, port))
        return True
    except OSError:
        print("Can't connect to PLC")
        sleep(3)
        print("Reconnecting....")
        return False

def readdata(data):
    a = 'RD '
    c = '\x0D'
    d = a+ data +c
    datasend = d.encode("UTF-8")
    soc.sendall(datasend)
    data = soc.recv(1024)
    datadeco = data.decode("UTF-8")
    data1 = int(datadeco)

    return data1

#Write data
def writedata(register, data):
    a = 'WR '
    b = ' '
    c = '\x0D'
    d = a+ register + b + str(data) + c
    datasend  = d.encode("UTF-8")
    soc.sendall(datasend)
    datares = soc.recv(1024)
    #print(datares)





def removefile():
    directory1 = 'C:/FH/camera1/'
    directory2 = 'C:/FH/camera2/'
    directory3 = 'C:/FH/camera3/'
    directory4 = 'C:/FH/camera4/'
    if not os.path.isdir(directory1):
        os.mkdir(directory1)
    if not os.path.isdir(directory2):
        os.mkdir(directory2)
    if not os.path.isdir(directory3):
        os.mkdir(directory3)
    if not os.path.isdir(directory4):
        os.mkdir(directory4)
    if os.listdir(directory1) != []:
        for i in glob.glob(directory1+'*'):
            for j in glob.glob(i+'/*'):
                os.remove(j)
            os.rmdir(i)

    if os.listdir(directory2) != []:
        for i in glob.glob(directory2+'*'):
            for j in glob.glob(i+'/*'):
                os.remove(j)
            os.rmdir(i)

    if os.listdir(directory3) != []:
        for i in glob.glob(directory3+'*'):
            for j in glob.glob(i+'/*'):
                os.remove(j)
            os.rmdir(i)

    if os.listdir(directory4) != []:
        for i in glob.glob(directory4+'*'):
            for j in glob.glob(i+'/*'):
                os.remove(j)
            os.rmdir(i)
    print('already delete folder')



def time_to_name():
    current_time = datetime.datetime.now() 
    name_folder = str(current_time)
    name_folder = list(name_folder)
    for i in range(len(name_folder)):
        if name_folder[i] == ':':
            name_folder[i] = '-'
        if name_folder[i] == ' ':
            name_folder[i] ='_'
        if name_folder[i] == '.':
            name_folder[i] ='-'
    name_folder = ''.join(name_folder)
    return name_folder




def load_theme():
    name_themes = []
    with open('static/theme.txt') as lines:
        for line in lines:
            _, name_theme = line.strip().split(':')
            name_themes.append(name_theme)
    return name_themes

def load_choosemodel():
    with open('static/choose_model.txt') as lines:
        for line in lines:
            _, name_model = line.strip().split('=')
    return name_model

def save_theme(name_theme):
    line = 'theme:' + name_theme
    with open('static/theme.txt','w') as f:
        f.write(line)


def save_choosemodel(name_model):
    line = 'choose_model=' + name_model
    with open('static/choose_model.txt','w') as f:
        f.write(line)

def load_model(i):
    with open('static/model'+ str(i) + '.txt','r') as lines:
        for line in lines:
            _, name_model = line.strip().split('=')
    return name_model

def save_model(i,name_model):
    line = 'model' + str(i) + '=' + name_model
    with open('static/model' + str(i) + '.txt','w') as f:
        f.write(line)


def str2bool(v):
    return v.lower() in ("yes", "true", "t", "1")

def load_all(model,i):
    values_all = []
    with open('static/all'+ str(i) + '.txt','r') as lines:
        for line in lines:
            _, name_all = line.strip().split('=')
            values_all.append(name_all)
    window['file_weights' + str(i)].update(value=values_all[0])
    window['conf_thres' + str(i)].update(value=values_all[1])
    a=1
    for item in range(len(model.names)):
        window[f'{model.names[item]}_' + str(i)].update(value=str2bool(values_all[a+1]))
        window[f'{model.names[item]}_OK_' + str(i)].update(value=str2bool(values_all[a+2]))
        window[f'{model.names[item]}_Num_' + str(i)].update(value=str(values_all[a+3]))
        window[f'{model.names[item]}_NG_' + str(i)].update(value=str2bool(values_all[a+4]))
        window[f'{model.names[item]}_Wn_' + str(i)].update(value=str(values_all[a+5]))
        window[f'{model.names[item]}_Wx_' + str(i)].update(value=str(values_all[a+6]))
        window[f'{model.names[item]}_Hn_' + str(i)].update(value=str(values_all[a+7]))
        window[f'{model.names[item]}_Hx_' + str(i)].update(value=str(values_all[a+8]))
        a += 8


def save_all(model,i):
    with open('static/all'+ str(i) + '.txt','w') as f:
        f.write('weights' + str(i) + '=' + str(values['file_weights' + str(i)]))
        f.write('\n')
        f.write('conf' + str(i) + '=' + str(values['conf_thres' + str(i)]))
        f.write('\n')

        for item in range(len(model.names)):
            f.write(str(f'{model.names[item]}_' + str(i)) + '=' + str(values[f'{model.names[item]}_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_OK_' + str(i)) + '=' + str(values[f'{model.names[item]}_OK_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Num_' + str(i)) + '=' + str(values[f'{model.names[item]}_Num_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_NG_' + str(i)) + '=' + str(values[f'{model.names[item]}_NG_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Wn_' + str(i)) + '=' + str(values[f'{model.names[item]}_Wn_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Wx_' + str(i)) + '=' + str(values[f'{model.names[item]}_Wx_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Hn_' + str(i)) + '=' + str(values[f'{model.names[item]}_Hn_' + str(i)]))
            f.write('\n')
            f.write(str(f'{model.names[item]}_Hx_' + str(i)) + '=' + str(values[f'{model.names[item]}_Hx_' + str(i)]))
            if item != len(model.names)-1:
                f.write('\n')


def load_all_sql(i,choose_model):
    conn = sqlite3.connect('modeldb_4_PLC_conf.db')
    cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,OK_Cam4,NG_Cam1,NG_Cam2,NG_Cam3,NG_Cam4,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_OK_Cam4,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Folder_NG_Cam4,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
    for row in cursor:
        #if row[0] == values['choose_model']:
        if row[0] == choose_model:
            row1_a, row1_b = row[1].strip().split('_')
            if row1_a == str(i) and row1_b == '0':
                window['file_weights' + str(i)].update(value=row[2])
                window['conf_thres' + str(i)].update(value=row[3])
                window['have_save_OK_1'].update(value=str2bool(row[4]))
                window['have_save_OK_2'].update(value=str2bool(row[5]))
                window['have_save_OK_3'].update(value=str2bool(row[6]))
                window['have_save_OK_4'].update(value=str2bool(row[7]))
                window['have_save_NG_1'].update(value=str2bool(row[8]))
                window['have_save_NG_2'].update(value=str2bool(row[9]))
                window['have_save_NG_3'].update(value=str2bool(row[10]))
                window['have_save_NG_4'].update(value=str2bool(row[11]))

                window['save_OK_1'].update(value=row[12])
                window['save_OK_2'].update(value=row[13])
                window['save_OK_3'].update(value=row[14])
                window['save_OK_4'].update(value=row[15])
                window['save_NG_1'].update(value=row[16])
                window['save_NG_2'].update(value=row[17])
                window['save_NG_3'].update(value=row[18])
                window['save_NG_4'].update(value=row[19])

                model = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)
            if row1_a == str(i):
                for item in range(len(model.names)):
                    if int(row1_b) == item:
                        window[f'{model.names[item]}_' + str(i)].update(value=str2bool(row[20]))
                        window[f'{model.names[item]}_OK_' + str(i)].update(value=str2bool(row[21]))
                        window[f'{model.names[item]}_Num_' + str(i)].update(value=str(row[22]))
                        window[f'{model.names[item]}_NG_' + str(i)].update(value=str2bool(row[23]))
                        window[f'{model.names[item]}_Wn_' + str(i)].update(value=str(row[24]))
                        window[f'{model.names[item]}_Wx_' + str(i)].update(value=str(row[25]))
                        window[f'{model.names[item]}_Hn_' + str(i)].update(value=str(row[26]))
                        window[f'{model.names[item]}_Hx_' + str(i)].update(value=str(row[27]))
                        window[f'{model.names[item]}_PLC_' + str(i)].update(value=str(row[28]))
                        window[f'OK_PLC_' + str(i)].update(value=str(row[29]))
                        window[f'{model.names[item]}_Conf_' + str(i)].update(value=str(row[30]))

    conn.close()


def save_all_sql(model,i,choose_model):
    conn = sqlite3.connect('modeldb_4_PLC_conf.db')
    cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,OK_Cam4,NG_Cam1,NG_Cam2,NG_Cam3,NG_Cam4,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_OK_Cam4,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Folder_NG_Cam4,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
    update = 0 

    for row in cursor:
        if row[0] == choose_model:            
            row1_a, _ = row[1].strip().split('_')
            if row1_a == str(i):
                conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera LIKE ?)", (choose_model,str(i) + '%'))
                for item in range(len(model.names)):
                    #conn.execute("UPDATE MYMODEL SET ChooseModel = ? , Camera = ?, Weights = ?,Confidence = ?, Joined = ?, Ok = ?, Num = ?, NG = ?, WidthMin = ?, WidthMax = ?, HeightMin = ?, HeightMax = ? WHERE (ChooseModel = ? AND Camera = ?)",(str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]),int(values['conf_thres' + str(i)]), str(values[model.names[item] + '_' + str(i)]), str(values[model.names[item]+ '_OK_' + str(i)]), int(values[model.names[item]+ '_Num_' + str(i)]), str(values[model.names[item]+ '_NG_' + str(i)]), int(values[model.names[item] + '_Wn_' + str(i)]), int(values[model.names[item] + '_Wx_' + str(i)]), int(values[model.names[item]+ '_Hn_' + str(i)]), int(values[model.names[item] + '_Hx_' + str(i)]), choose_model,str(i) + '_' + str(item)))
                    #conn.execute("DELETE FROM MYMODEL WHERE (ChooseModel = ? AND Camera = ?)", (choose_model,str(i) + '_' + str(item)))
                    conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,OK_Cam4,NG_Cam1,NG_Cam2,NG_Cam3,NG_Cam4,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_OK_Cam4,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Folder_NG_Cam4,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf) \
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_OK_4']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['have_save_NG_4']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_OK_4']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values['save_NG_4']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]), int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]), int(values[f'{model.names[item]}_Conf_' + str(i)])))           
                    #conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax) \
                    #    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)])))           
                    update = 1
                break

    if update == 0:
        for item in range(len(model.names)):
            conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,OK_Cam4,NG_Cam1,NG_Cam2,NG_Cam3,NG_Cam4,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_OK_Cam4,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Folder_NG_Cam4,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax, PLC_NG,PLC_OK,Conf) \
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_OK_4']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['have_save_NG_4']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_OK_4']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values['save_NG_4']),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)]),int(values[f'{model.names[item]}_PLC_' + str(i)]), int(values['OK_PLC_' + str(i)]), int(values[f'{model.names[item]}_Conf_' + str(i)])))
            #conn.execute("INSERT INTO MYMODEL (ChooseModel,Camera, Weights,Confidence,Joined,Ok,Num,NG,WidthMin, WidthMax,HeightMin,HeightMax) \
            #    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", (str(values['choose_model']),str(i)+ '_' +str(item) ,str(values['file_weights' + str(i)]), int(values['conf_thres' + str(i)]),str(values[f'{model.names[item]}_' + str(i)]), str(values[f'{model.names[item]}_OK_' + str(i)]), int(values[f'{model.names[item]}_Num_' + str(i)]), str(values[f'{model.names[item]}_NG_' + str(i)]), int(values[f'{model.names[item]}_Wn_' + str(i)]), int(values[f'{model.names[item]}_Wx_' + str(i)]), int(values[f'{model.names[item]}_Hn_' + str(i)]), int(values[f'{model.names[item]}_Hx_' + str(i)])))
        
    for row in cursor:
        if row[0] == choose_model:
            conn.execute("UPDATE MYMODEL SET OK_Cam1 = ? , OK_Cam2 = ?,OK_Cam3 = ? , OK_Cam4 = ?, NG_Cam1 = ?,NG_Cam2 = ?, NG_Cam3 = ?,NG_Cam4 = ?, Folder_OK_Cam1 = ?, Folder_OK_Cam2 = ?,Folder_OK_Cam3 = ?,Folder_OK_Cam4 = ?, Folder_NG_Cam1 = ?, Folder_NG_Cam2 = ?,Folder_NG_Cam3 = ?, Folder_NG_Cam4 = ? WHERE ChooseModel = ? ",(str(values['have_save_OK_1']),str(values['have_save_OK_2']),str(values['have_save_OK_3']),str(values['have_save_OK_4']),str(values['have_save_NG_1']),str(values['have_save_NG_2']),str(values['have_save_NG_3']),str(values['have_save_NG_4']),str(values['save_OK_1']),str(values['save_OK_2']),str(values['save_OK_3']),str(values['save_OK_4']),str(values['save_NG_1']),str(values['save_NG_2']),str(values['save_NG_3']),str(values['save_NG_4']),choose_model))


    conn.commit()
    conn.close()






def program_camera1_FH(model,size,conf):
    read_1900 = readdata('DM1900') # doc thanh ghi 1900
    #print('Cam 1' , read_1900)
    if read_1900 == 1:  # gia tri 1
        directory1 = 'C:/FH/camera1/'
        if not os.path.isdir(directory1):
            os.mkdir(directory1)
        if os.listdir(directory1) == []:
            print('folder 1 empty')
        else:
            print('received folder 1')
            for filename1 in glob.glob('C:/FH/camera1/*'):
                for path1 in glob.glob(filename1 + '/*'):
                    name = path1[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img1_orgin = cv2.imread(path1)
                        while type(img1_orgin) == type(None):
                            print('loading img 1...')
                            for path1 in glob.glob(filename1 + '/*'):
                                img1_orgin = cv2.imread(path1)

                        img1_save = img1_orgin
                        #img1_orgin = cv2.resize(img1_orgin,(640,480))                   
                        print('CAM 1')
                        t1 = time.time()

                        # ghi vao D1900 gia tri 0
                        writedata('DM1900.U',0) 

                        img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)

                        result1 = model(img1_orgin,size= size,conf = conf) 
                        table1 = result1.pandas().xyxy[0]
                        area_remove1 = []

                        myresult1 =0 

                        for item in range(len(table1.index)):
                            width1 = table1['xmax'][item] - table1['xmin'][item]
                            height1 = table1['ymax'][item] - table1['ymin'][item]
                            conf1 = table1['confidence'][item] * 100

                            #area1 = width1*height1
                            label_name = table1['name'][item]
                            for i1 in range(len(model1.names)):
                                if values[f'{model1.names[i1]}_1'] == True:
                                    #if values[f'{model1.names[i1]}_WH'] == True:
                                    if label_name == model1.names[i1]:
                                        if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)
                                        elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                            table1.drop(item, axis=0, inplace=True)
                                            area_remove1.append(item)                                                
                                if values[f'{model1.names[i1]}_1'] == False:
                                    if label_name == model1.names[i1]:
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)

                        names1 = list(table1['name'])

                        show1 = np.squeeze(result1.render(area_remove1))
                        show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                        show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)

                        #ta = time.time()
                        for i1 in range(len(model1.names)):
                            register_ng = str(values[f'{model1.names[i1]}_PLC_1'])
                            if values[f'{model1.names[i1]}_OK_1'] == True:
                                len_name1 = 0
                                for name1 in names1:
                                    if name1 == model1.names[i1]:
                                        len_name1 +=1
                                if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam1'].update(value= 'NG', text_color='red')
                                    # if values['have_save_NG_1']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)
                                    myresult1 = 1
                                    #break

                            if values[f'{model1.names[i1]}_NG_1'] == True:
                                if model1.names[i1] in names1:
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam1'].update(value= 'NG', text_color='red')    
                                    # if values['have_save_NG_1']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)
                                    myresult1 = 1         
                                    #break    

                        if myresult1 == 0:
                            print('OK')
                            writedata('DM' + str(values['OK_PLC_1']) + '.U',1) 
                            t2 = time.time() - t1
                            print(t2) 
                            cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                            window['result_cam1'].update(value= 'OK', text_color='green')
                            if values['have_save_OK_1']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_OK_1']  + '/' + name_folder_ng + '.jpg',img1_save)
                        else:
                            if values['have_save_NG_1']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_NG_1']  + '/' + name_folder_ng + '.jpg',img1_save)

                        writedata('DM1910.U',1) 

                        time_cam1 = str(int(t2*1000)) + 'ms'
                        window['time_cam1'].update(value= time_cam1, text_color='black') 
                    

                        imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                        window['image1'].update(data= imgbytes1)
                        #print('---------------------------------------------')

                    if os.path.isfile(path1):
                        os.remove(path1)
                while os.path.isdir(directory1):
                    try:
                        shutil.rmtree(directory1)
                    except:
                        print('Error delete folder 1')



  
def program_camera2_FH(model,size,conf):
    read_1902 = readdata('DM1902') # doc thanh ghi 1902
    #print('Cam 2' , read_1902)
    if read_1902 == 1:  # gia tri 1
        directory2 = 'C:/FH/camera2/'
        if not os.path.isdir(directory2):
            os.mkdir(directory2)
        if os.listdir(directory2) == []:
            print('folder 2 empty')
        else:
            print('received folder 2')

            for filename2 in glob.glob('C:/FH/camera2/*'):
                for path2 in glob.glob(filename2 + '/*'):
                    name = path2[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img2_orgin = cv2.imread(path2)
                        while type(img2_orgin) == type(None):
                            print('loading img 2...')
                            for path2 in glob.glob(filename2 + '/*'):
                                img2_orgin = cv2.imread(path2)

                        #img2_orgin = cv2.resize(img2_orgin,(640,480))
                        img2_save = img2_orgin
                        #edit
                        print('CAM 2')
                        t1 = time.time()

                        # ghi vao D1902 gia tri 0
                        writedata('DM1902.U',0) 


                        img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)

                        result2 = model(img2_orgin,size= size,conf = conf) 
                        table2 = result2.pandas().xyxy[0]
                        area_remove2 = []

                        myresult2 =0 

                        for item in range(len(table2.index)):
                            width2 = table2['xmax'][item] - table2['xmin'][item]
                            height2 = table2['ymax'][item] - table2['ymin'][item]
                            conf2 = table2['confidence'][item] * 100

                            label_name = table2['name'][item]
                            for i2 in range(len(model2.names)):
                                if values[f'{model2.names[i2]}_2'] == True:
                                    if label_name == model2.names[i2]:
                                        if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item)
                                        elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                            table2.drop(item, axis=0, inplace=True)
                                            area_remove2.append(item) 

                                if values[f'{model2.names[i2]}_2'] == False:
                                    if label_name == model2.names[i2]:
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)

                        names2 = list(table2['name'])

                        show2 = np.squeeze(result2.render(area_remove2))
                        show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                        show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
                        #ta = time.time()
                        for i2 in range(len(model2.names)):
                            register_ng = str(values[f'{model2.names[i2]}_PLC_2'])
                            if values[f'{model2.names[i2]}_OK_2'] == True:
                                len_name2 = 0
                                for name2 in names2:
                                    if name2 == model2.names[i2]:
                                        len_name2 +=1
                                if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam2'].update(value= 'NG', text_color='red')
                                    # if values['have_save_NG_2']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)
                                    myresult2 = 1
                                    #break

                            if values[f'{model2.names[i2]}_NG_2'] == True:
                                if model2.names[i2] in names2:
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam2'].update(value= 'NG', text_color='red')    
                                    # if values['have_save_NG_2']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)
                                    myresult2 = 1         
                                    #break    

                        if myresult2 == 0:
                            print('OK')
                            writedata('DM' + str(values['OK_PLC_2']) + '.U',1) 
                            t2 = time.time() - t1
                            print(t2) 
                            cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                            window['result_cam2'].update(value= 'OK', text_color='green')
                            if values['have_save_OK_2']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_OK_2']  + '/' + name_folder_ng + '.jpg',img2_save)
                        else:
                            if values['have_save_NG_2']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_NG_2']  + '/' + name_folder_ng + '.jpg',img2_save)

                        # ghi vao D1912 gia tri 1
                        writedata('DM1912.U',1) 

                        time_cam2 = str(int(t2*1000)) + 'ms'
                        window['time_cam2'].update(value= time_cam2, text_color='black') 
                    

                        imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                        window['image2'].update(data= imgbytes2)
                        print('---------------------------------------------')
                    if os.path.isfile(path2):
                        os.remove(path2)
                while os.path.isdir(directory2):
                    try:
                        shutil.rmtree(directory2)
                        
                    except:
                        print('Error delete folder 2')



def program_camera3_FH(model,size,conf):
    read_1904 = readdata('DM1904') # doc thanh ghi 1904
    #print('Cam 3' , read_1904)
    if read_1904 == 1:  # gia tri 1
        directory3 = 'C:/FH/camera3/'
        if not os.path.isdir(directory3):
            os.mkdir(directory3)
        if os.listdir(directory3) == []:
            print('folder 3 empty')
        else:
            print('received folder 3')

            for filename3 in glob.glob('C:/FH/camera3/*'):
                for path3 in glob.glob(filename3 + '/*'):
                    name = path3[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img3_orgin = cv2.imread(path3)
                        while type(img3_orgin) == type(None):
                            print('loading img 3...')
                            for path3 in glob.glob(filename3 + '/*'):
                                img3_orgin = cv2.imread(path3)

                        img3_save = img3_orgin
                        #img3_orgin = cv2.resize(img3_orgin,(640,480))

                        print('CAM 3')
                        t1 = time.time()

                        # ghi vao D1904 gia tri 0
                        writedata('DM1904.U',0) 


                        img3_orgin = cv2.cvtColor(img3_orgin, cv2.COLOR_BGR2RGB)

                        result3 = model(img3_orgin,size= size,conf = conf) 
                        table3 = result3.pandas().xyxy[0]
                        area_remove3 = []

                        myresult3 =0 

                        for item in range(len(table3.index)):
                            width3 = table3['xmax'][item] - table3['xmin'][item]
                            height3 = table3['ymax'][item] - table3['ymin'][item]
                            conf3 = table3['confidence'][item] * 100

                            label_name = table3['name'][item]
                            for i3 in range(len(model3.names)):
                                if values[f'{model3.names[i3]}_3'] == True:
                                    #if values[f'{model3.names[i3]}_WH'] == True:
                                    if label_name == model3.names[i3]:
                                        if width3 < int(values[f'{model3.names[i3]}_Wn_3']): 
                                            table3.drop(item, axis=0, inplace=True)
                                            area_remove3.append(item)
                                        elif width3 > int(values[f'{model3.names[i3]}_Wx_3']): 
                                            table3.drop(item, axis=0, inplace=True)
                                            area_remove3.append(item)
                                        elif height3 < int(values[f'{model3.names[i3]}_Hn_3']): 
                                            table3.drop(item, axis=0, inplace=True)
                                            area_remove3.append(item)
                                        elif height3 > int(values[f'{model3.names[i3]}_Hx_3']): 
                                            table3.drop(item, axis=0, inplace=True)
                                            area_remove3.append(item)
                                        elif conf3 < int(values[f'{model3.names[i3]}_Conf_3']):
                                            table3.drop(item, axis=0, inplace=True)
                                            area_remove3.append(item) 

                                if values[f'{model3.names[i3]}_3'] == False:
                                    if label_name == model3.names[i3]:
                                        table3.drop(item, axis=0, inplace=True)
                                        area_remove3.append(item)

                        names3 = list(table3['name'])

                        show3 = np.squeeze(result3.render(area_remove3))
                        show3 = cv2.resize(show3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                        show3 = cv2.cvtColor(show3, cv2.COLOR_BGR2RGB)
                        #ta = time.time()
                        for i3 in range(len(model3.names)):
                            register_ng = str(values[f'{model3.names[i3]}_PLC_3'])
                            if values[f'{model3.names[i3]}_OK_3'] == True:
                                len_name3 = 0
                                for name3 in names3:
                                    if name3 == model3.names[i3]:
                                        len_name3 +=1
                                if len_name3 != int(values[f'{model3.names[i3]}_Num_3']):
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam3'].update(value= 'NG', text_color='red')
                                    # if values['have_save_NG_3']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_3']  + '/' + name_folder_ng + '.jpg',img3_save)
                                    myresult3 = 1
                                    #break

                            if values[f'{model3.names[i3]}_NG_3'] == True:
                                if model3.names[i3] in names3:
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam3'].update(value= 'NG', text_color='red')    
                                    # if values['have_save_NG_3']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_3']  + '/' + name_folder_ng + '.jpg',img3_save)
                                    myresult3 = 1         
                                    #break    

                        if myresult3 == 0:
                            print('OK')
                            writedata('DM' + str(values['OK_PLC_3']) + '.U',1) 
                            t2 = time.time() - t1
                            print(t2) 
                            cv2.putText(show3, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                            window['result_cam3'].update(value= 'OK', text_color='green')
                            if values['have_save_OK_3']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_OK_3']  + '/' + name_folder_ng + '.jpg',img3_save)
                        else:
                            if values['have_save_NG_3']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_NG_3']  + '/' + name_folder_ng + '.jpg',img3_save)

                        # ghi vao D1914 gia tri 1
                        writedata('DM1914.U',1) 

                        time_cam3 = str(int(t2*1000)) + 'ms'
                        window['time_cam3'].update(value= time_cam3, text_color='black') 
                    

                        imgbytes3 = cv2.imencode('.png',show3)[1].tobytes()
                        window['image3'].update(data= imgbytes3)
                        print('---------------------------------------------')
                    if os.path.isfile(path3):
                        os.remove(path3)
                while os.path.isdir(directory3):
                    try:
                        shutil.rmtree(directory3)
                    except:
                        print('Error delete folder 3')



def program_camera4_FH(model,size,conf):
    read_1906 = readdata('DM1906') # doc thanh ghi 1906
    #print('Cam 4' , read_1906)
    if read_1906 == 1:  # gia tri 1
        directory4 = 'C:/FH/camera4/'
        if not os.path.isdir(directory4):
            os.mkdir(directory4)
        if os.listdir(directory4) == []:
            print('folder 4 empty')
        else:
            print('received folder 4')

            for filename4 in glob.glob('C:/FH/camera4/*'):
                for path4 in glob.glob(filename4 + '/*'):
                    name = path4[-18:]
                    if name == 'Input0_Camera0.jpg':
                        img4_orgin = cv2.imread(path4)
                        while type(img4_orgin) == type(None):
                            print('loading img 4...')
                            for path4 in glob.glob(filename4 + '/*'):
                                img4_orgin = cv2.imread(path4)

                        #img4_orgin = cv2.resize(img4_orgin,(640,480))

                        img4_save = img4_orgin
                        print('CAM 4')
                        t1 = time.time()

                        # ghi vao D1906 gia tri 0
                        writedata('DM1906.U',0) 

                        img4_orgin = cv2.cvtColor(img4_orgin, cv2.COLOR_BGR2RGB)

                        result4 = model(img4_orgin,size= size,conf = conf) 
                        table4 = result4.pandas().xyxy[0]
                        area_remove4 = []

                        myresult4 =0 

                        for item in range(len(table4.index)):
                            width4 = table4['xmax'][item] - table4['xmin'][item]
                            height4 = table4['ymax'][item] - table4['ymin'][item]
                            label_name = table4['name'][item]
                            conf4 = table4['confidence'][item] * 100

                            for i4 in range(len(model4.names)):
                                if values[f'{model4.names[i4]}_4'] == True:
                                    if label_name == model4.names[i4]:
                                        if width4 < int(values[f'{model4.names[i4]}_Wn_4']): 
                                            table4.drop(item, axis=0, inplace=True)
                                            area_remove4.append(item)
                                        elif width4 > int(values[f'{model4.names[i4]}_Wx_4']): 
                                            table4.drop(item, axis=0, inplace=True)
                                            area_remove4.append(item)
                                        elif height4 < int(values[f'{model4.names[i4]}_Hn_4']): 
                                            table4.drop(item, axis=0, inplace=True)
                                            area_remove4.append(item)
                                        elif height4 > int(values[f'{model4.names[i4]}_Hx_4']): 
                                            table4.drop(item, axis=0, inplace=True)
                                            area_remove4.append(item)
                                        elif conf4 < int(values[f'{model4.names[i4]}_Conf_4']):
                                            table4.drop(item, axis=0, inplace=True)
                                            area_remove4.append(item) 

                                if values[f'{model4.names[i4]}_4'] == False:
                                    if label_name == model4.names[i4]:
                                        table4.drop(item, axis=0, inplace=True)
                                        area_remove4.append(item)

                        names4 = list(table4['name'])

                        show4 = np.squeeze(result4.render(area_remove4))
                        show4 = cv2.resize(show4, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                        show4 = cv2.cvtColor(show4, cv2.COLOR_BGR2RGB)
                        #ta = time.time()
                        for i4 in range(len(model4.names)):
                            register_ng = str(values[f'{model4.names[i4]}_PLC_4'])
                            if values[f'{model4.names[i4]}_OK_4'] == True:
                                len_name4 = 0
                                for name4 in names4:
                                    if name4 == model4.names[i4]:
                                        len_name4 +=1
                                if len_name4 != int(values[f'{model4.names[i4]}_Num_4']):
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam4'].update(value= 'NG', text_color='red')
                                    # if values['have_save_NG_4']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_4']  + '/' + name_folder_ng + '.jpg',img4_save)
                                    myresult4 = 1
                                    #break

                            if values[f'{model4.names[i4]}_NG_4'] == True:
                                if model4.names[i4] in names4:
                                    print('NG')
                                    writedata('DM' + register_ng + '.U',1) 
                                    t2 = time.time() - t1
                                    print(t2) 
                                    cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                    window['result_cam4'].update(value= 'NG', text_color='red')    
                                    # if values['have_save_NG_4']:
                                    #     name_folder_ng = time_to_name()
                                    #     cv2.imwrite(values['save_NG_4']  + '/' + name_folder_ng + '.jpg',img4_save)
                                    myresult4 = 1         
                                    #break    

                        if myresult4 == 0:
                            print('OK')
                            writedata('DM' + str(values['OK_PLC_4']) + '.U',1) 
                            t2 = time.time() - t1
                            print(t2) 
                            cv2.putText(show4, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                            window['result_cam4'].update(value= 'OK', text_color='green')
                            if values['have_save_OK_4']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_OK_4']  + '/' + name_folder_ng + '.jpg',img4_save)

                        else:
                            if values['have_save_NG_4']:
                                name_folder_ng = time_to_name()
                                cv2.imwrite(values['save_NG_4']  + '/' + name_folder_ng + '.jpg',img4_save)

                        # ghi vao D1916 gia tri 1
                        writedata('DM1916.U',1) 

                        time_cam4 = str(int(t2*1000)) + 'ms'
                        window['time_cam4'].update(value= time_cam4, text_color='black') 
                    

                        imgbytes4 = cv2.imencode('.png',show4)[1].tobytes()
                        window['image4'].update(data= imgbytes4)
                        print('---------------------------------------------')
                    if os.path.isfile(path4):
                        os.remove(path4)
                # while os.path.isdir(filename4):
                #     try:
                #         shutil.rmtree(filename4)
                while os.path.isdir(directory4):
                    try:
                        shutil.rmtree(directory4)
                    except:
                        print('Error delete folder 4')


def excel_pp(var_plc):
    wb = openpyxl.load_workbook("excel/ppbydate.xlsx")
    ws = wb.active

    now = datetime.datetime.now()
    col1 = now.strftime("%d/%m/%Y %H:%M:%S")

    phe_pham = ["CUON CAM", "CACBON TAY CHOI","HAN CHOI","HAN CHAU","DE VO NHO","TU DIEN", "CONG CHAU DIEN","BUI CHI"]
    all_cam = ["CAM 1","CAM 2","CAM 3","CAM 4"]

    tt = 0
    for pp in range(len(phe_pham)):
        for al in range(len(all_cam)):
            if tt == var_plc:
                col2 = all_cam[al] + " " + phe_pham[pp]
            tt+=1

    ws.append([col1, col2])

    for row in range(ws.max_row, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            #print(col[row].value)
            d = ws.cell(row = row, column = col)
            #currentCell = ws.cell(col[row])
            d.alignment = Alignment(horizontal='center')
            #d.style.alignment.horizontal = 'center'
            d.font = Font(name= 'Calibri', size=12)

    wb.save("excel/ppbydate.xlsx")
    try:
        shutil.copy("excel/ppbydate.xlsx", "C:/Users/Administrator/Desktop/excel/ppbydate.xlsx")
    except:
        pass



def excel_sang():
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:M1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:M2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Cuộn cảm'
    HomNay['E3'] = 'Cacbon tay chổi'
    HomNay['F3'] = 'Hàn chổi'
    HomNay['G3'] = 'Hàn chấu '
    HomNay['H3'] = 'Đế vỏ nhỏ'
    HomNay['I3'] = 'Tụ điện'
    HomNay['J3'] = 'Cong chấu điện'
    HomNay['K3'] = 'Bụi chì'
    HomNay['L3'] = 'PP khác (nhiều hạng mục)'
    HomNay['M3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

    HomNay.column_dimensions['A'].width = 20
    HomNay.column_dimensions['B'].width = 20
    HomNay.column_dimensions['C'].width = 25
    HomNay.column_dimensions['D'].width = 18
    HomNay.column_dimensions['E'].width = 18
    HomNay.column_dimensions['F'].width = 18
    HomNay.column_dimensions['G'].width = 18
    HomNay.column_dimensions['H'].width = 18
    HomNay.column_dimensions['I'].width = 18
    HomNay.column_dimensions['J'].width = 18
    HomNay.column_dimensions['K'].width = 18
    HomNay.column_dimensions['L'].width = 25
    HomNay.column_dimensions['M'].width = 25

    wb.remove(wb['Sheet'])

    wb.save(f"excel/{mydate}_Ngay.xlsx")
    try:
        shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/Users/Administrator/Desktop/excel/{mydate}_Ngay.xlsx")
    except:
        pass
    writedata('DM6050.U',0) 

def excel_dem():
    today = date.today()
    mydate = today.strftime("%Y_%m_%d")
    wb = openpyxl.Workbook()

    HomNay = wb.create_sheet("Data")


    HomNay.merge_cells('A1:M1')
    HomNay.merge_cells('A2:A3')
    HomNay.merge_cells('B2:B3')
    HomNay.merge_cells('C2:M2')
    #HomNay.unmerge_cells('A2:D2')
    HomNay['A1'] = 'DỮ LIỆU MÁY NQVNHT RS656 A17'
    HomNay['A1'].alignment = Alignment(horizontal='center')
    HomNay['A1'].font = Font(name= 'Calibri', size=20)

    HomNay['A2'] = 'Ngày sản xuất'
    HomNay['A2'].alignment = Alignment(horizontal='center')
    HomNay['A2'].font = Font(name= 'Calibri', size=12)
    HomNay['B2'] = 'Giờ lưu dữ liệu'
    HomNay['B2'].alignment = Alignment(horizontal='center')
    HomNay['B2'].font = Font(name= 'Calibri', size=12)
    HomNay['C2'] = 'HẠNG MỤC PHẾ PHẨM'
    HomNay['C2'].alignment = Alignment(horizontal='center')
    HomNay['C2'].font = Font(name= 'Calibri', size=12)
    HomNay['C3'] = 'Tổng số lượng sản xuất'
    HomNay['D3'] = 'Cuộn cảm'
    HomNay['E3'] = 'Cacbon tay chổi'
    HomNay['F3'] = 'Hàn chổi'
    HomNay['G3'] = 'Hàn chấu '
    HomNay['H3'] = 'Đế vỏ nhỏ'
    HomNay['I3'] = 'Tụ điện'
    HomNay['J3'] = 'Cong chấu điện'
    HomNay['K3'] = 'Bụi chì'
    HomNay['L3'] = 'PP khác (nhiều hạng mục)'
    HomNay['M3'] = 'Tổng số lượng PP'

    for i in range(67,77):
        HomNay[f'{str(chr(i))}3'].alignment = Alignment(horizontal='center')
        HomNay[f'{str(chr(i))}3'].font = Font(name= 'Calibri', size=12) 

    HomNay.column_dimensions['A'].width = 20
    HomNay.column_dimensions['B'].width = 20
    HomNay.column_dimensions['C'].width = 25
    HomNay.column_dimensions['D'].width = 18
    HomNay.column_dimensions['E'].width = 18
    HomNay.column_dimensions['F'].width = 18
    HomNay.column_dimensions['G'].width = 18
    HomNay.column_dimensions['H'].width = 18
    HomNay.column_dimensions['I'].width = 18
    HomNay.column_dimensions['J'].width = 18
    HomNay.column_dimensions['K'].width = 18
    HomNay.column_dimensions['L'].width = 25
    HomNay.column_dimensions['M'].width = 25

    wb.remove(wb['Sheet'])

    wb.save(f"excel/{mydate}_Dem.xlsx")
    try:
        shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/Users/Administrator/Desktop/excel/{mydate}_Dem.xlsx")
    except:
        pass
    writedata('DM6050.U',0) 
                        
def make_window(theme):
    sg.theme(theme)

    #file_img = [("JPEG (*.jpg)",("*jpg","*.png"))]

    file_weights = [('Weights (*.pt)', ('*.pt'))]

    # menu = [['Application', ['Connect PLC','Interrupt Connect PLC','Exit']],
    #         ['Tool', ['Check Cam','Change Theme']],
    #         ['Help',['About']]]

    right_click_menu = [[], ['Exit','Administrator','Change Theme']]


    layout_main = [

        [
        sg.Text('CAM 1',justification='center' ,font= ('Helvetica',int(30/1.4)),text_color='red',expand_x=True),
        sg.Text('CAM 2',justification='center' ,font= ('Helvetica',int(30/1.4)),text_color='red', expand_x=True),
        ],

        [

        #1
        sg.Frame('',[
            [sg.Image(filename='', size=(image_width_display,image_height_display),key='image1',background_color='black'),

            sg.Frame('',[
                [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam1'),sg.Text(' '), sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change1')],
                [sg.Text('')],
                [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Stop1'), sg.Text(' '),sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Pic1')],
                [sg.Text('')],
                [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Snap1'), sg.Text(' '),sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Detect1')],
                [sg.Text('')],
                [sg.Checkbox('Check',size=(6,1),font=('Helvetica',14), key='check_model1',enable_events=True,expand_x=True, expand_y=True),sg.Text(' '), sg.Combo(values=['1','2','3','4','5','6','7','8','9'], default_value='1',font=('Helvetica',20),size=(5, 100),text_color='navy',enable_events= True, key='choose_model')],
                [sg.Text('',font=('Helvetica',int(70/1.4)), justification='center', key='result_cam1',expand_x=True)],
                [sg.Text('',font=('Helvetica',int(25/1.4)), justification='center', key='time_cam1', expand_x=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],           
        ]),

        #2
        sg.Frame('',[
            [sg.Image(filename='', size=(image_width_display,image_height_display),key='image2',background_color='black'),

            sg.Frame('',[
                [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam2'),sg.Text(' '), sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change2')],
                [sg.Text('')],
                [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Stop2'), sg.Text(' '),sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Pic2')],
                [sg.Text('')],
                [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Snap2'), sg.Text(' '),sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Detect2')],
                [sg.Text('')],
                [sg.Checkbox('Check',size=(6,1),font=('Helvetica',14), key='check_model2',enable_events=True,expand_x=True, expand_y=True),sg.Text(' '*25)],
                [sg.Text('',font=('Helvetica',int(70/1.4)), justification='center', key='result_cam2',expand_x=True)],
                [sg.Text('',font=('Helvetica',int(25/1.4)), justification='center', key='time_cam2', expand_x=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],           
        ]),
        ],

        [
        sg.Text('CAM 3',justification='center' ,font= ('Helvetica',int(30/1.4)),text_color='red',expand_x=True),
        sg.Text('CAM 4',justification='center' ,font= ('Helvetica',int(30/1.4)),text_color='red', expand_y=True),
        ],
    
        [

        #3
        sg.Frame('',[
            [sg.Image(filename='', size=(image_width_display,image_height_display),key='image3',background_color='black'),

            sg.Frame('',[
                [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam3'),sg.Text(' '), sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change3')],
                [sg.Text('')],
                [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Stop3'), sg.Text(' '),sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Pic3')],
                [sg.Text('')],
                [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Snap3'), sg.Text(' '),sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Detect3')],
                [sg.Text('')],
                [sg.Checkbox('Check',size=(6,1),font=('Helvetica',14), key='check_model3',enable_events=True,expand_x=True, expand_y=True),sg.Text(' '*25)],
                [sg.Text('',font=('Helvetica',int(70/1.4)), justification='center', key='result_cam3',expand_x=True)],
                [sg.Text('',font=('Helvetica',int(25/1.4)), justification='center', key='time_cam3', expand_x=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],           
        ]),

        #4
        sg.Frame('',[
            [sg.Image(filename='', size=(image_width_display,image_height_display),key='image4',background_color='black'),

            sg.Frame('',[
                [sg.Button('Webcam', size=(8,1),  font=('Helvetica',14),disabled=True ,key= 'Webcam4'),sg.Text(' '), sg.Button('Change', size=(8,1), font=('Helvetica',14), disabled= True, key= 'Change4')],
                [sg.Text('')],
                [sg.Button('Stop', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Stop4'), sg.Text(' '),sg.Button('Pic', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Pic4')],
                [sg.Text('')],
                [sg.Button('Snap', size=(8,1), font=('Helvetica',14),disabled=True ,key= 'Snap4'), sg.Text(' '),sg.Button('Detect', size=(8,1), font=('Helvetica',14),disabled=True,key= 'Detect4')],
                [sg.Text('')],
                [sg.Checkbox('Check',size=(6,1),font=('Helvetica',14), key='check_model4',enable_events=True,expand_x=True, expand_y=True),sg.Text(' '*25)],
                [sg.Text('',font=('Helvetica',int(70/1.4)), justification='center', key='result_cam4',expand_x=True)],
                [sg.Text('',font=('Helvetica',int(25/1.4)), justification='center', key='time_cam4', expand_x=True)],
                ],element_justification='center', vertical_alignment='top', relief= sg.RELIEF_FLAT),
            ],           
        ]),

        ],
    
    ] 

    layout_option1a = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            #[sg.Text('Location', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='location_weights1',readonly= True, text_color='navy',enable_events= True),
            #sg.FolderBrowse(size=(15,1), font=('Helvetica',10),key= 'folder_browse1',enable_events=True)],
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights1',readonly= True, text_color='navy',enable_events= True),
            #[sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Combo(values='', font=('Helvetica',12),size=(59, 30),text_color='navy',enable_events= True, key='file_weights1'),],
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse1',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_1')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres1'),]
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],
        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model1.names[i1]}_1',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model1.names[i1]}_1',enable_events=True, disabled=True), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_OK_1',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Num_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_NG_1',enable_events=True, disabled=True), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_PLC_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model1.names[i1]}_Conf_1',enable_events=True, disabled=True),  
            ] for i1 in range(len(model1.names)-10)
        ], relief=sg.RELIEF_FLAT)],
        
        ])]
    ]
    
    layout_option1b = [
        [sg.Frame('',[

        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],

        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model1.names[i1]}_1',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model1.names[i1]}_1',enable_events=True, disabled=True), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_OK_1',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Num_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model1.names[i1]}_NG_1',enable_events=True, disabled=True), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Wx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hn_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_Hx_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model1.names[i1]}_PLC_1',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model1.names[i1]}_Conf_1',enable_events=True, disabled=True),            
                        
            ] for i1 in range(len(model1.names)-10,len(model1.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_1',text_color='navy',enable_events=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData1',enable_events=True)] 
        ])]
    ]
       

    layout_option2_1 = [
        [sg.Frame('',[
        [sg.Frame('',[
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights2',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse2',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_2')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres2')],

        ], relief=sg.RELIEF_FLAT, expand_y= True),],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],

        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model2.names[i2]}_2',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model2.names[i2]}_2',enable_events=True, disabled=True), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model2.names[i2]}_OK_2',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Num_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model2.names[i2]}_NG_2',enable_events=True, disabled=True), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Wn_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Wx_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Hn_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_Hx_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model2.names[i2]}_PLC_2',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model2.names[i2]}_Conf_2',enable_events=True, disabled=True),            
            ] for i2 in range(len(model2.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_2',text_color='navy',enable_events=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData2',enable_events=True)] 
        ])]
    ]

    layout_option2 = [[sg.Column(layout_option2_1,scrollable = True)]]
    layout_option3_1 = [
        [sg.Frame('',[
        [sg.Frame('',
        [   
            #[sg.Text('Location', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='location_weights4',readonly= True, text_color='navy',enable_events= True),
            #sg.FolderBrowse(size=(15,1), font=('Helvetica',10),key= 'folder_browse3',enable_events=True)],
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights3',readonly= True, text_color='navy',enable_events= True),
            #[sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Combo(values='', font=('Helvetica',12),size=(59, 30),text_color='navy',enable_events= True, key='file_weights3'),],
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse3',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_3')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres3'),]
        ], relief=sg.RELIEF_FLAT),
        ],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],

        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model3.names[i3]}_3',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model3.names[i3]}_3',enable_events=True, disabled=True), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model3.names[i3]}_OK_3',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Num_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model3.names[i3]}_NG_3',enable_events=True, disabled=True), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Wn_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Wx_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Hn_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_Hx_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model3.names[i3]}_PLC_3',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model3.names[i3]}_Conf_3',enable_events=True, disabled=True),    
            ] for i3 in range(len(model3.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_3',text_color='navy',enable_events=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData3',enable_events=True)] 
        ])]
    ]
    
    layout_option3 = [[sg.Column(layout_option3_1,scrollable = True)]]


    layout_option4 = [
        [sg.Frame('',[
        [sg.Frame('',[
            [sg.Text('Weights', size=(12,1), font=('Helvetica',15),text_color='red'), sg.Input(size=(60,1), font=('Helvetica',12), key='file_weights4',readonly= True, text_color='navy',enable_events= True),
            sg.Frame('',[
                [sg.FileBrowse(file_types= file_weights, size=(12,1), font=('Helvetica',10),key= 'file_browse4',enable_events=True, disabled=True)]
            ], relief= sg.RELIEF_FLAT),
            sg.Frame('',[
                [sg.Button('Change Model', size=(14,1), font=('Helvetica',10), disabled= True, key= 'Change_4')]
            ], relief= sg.RELIEF_FLAT),],
            [sg.Text('Confidence',size=(12,1),font=('Helvetica',15), text_color='red'), sg.Slider(range=(1,100),orientation='h',size=(60,20),font=('Helvetica',11),disabled=True, key= 'conf_thres4')],

        ], relief=sg.RELIEF_FLAT, expand_y= True),],
        [sg.Frame('',[
            [sg.Text('Name',size=(15,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Join',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('OK',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Num',size=(7,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('NG',size=(8,1),font=('Helvetica',15), text_color='red'),  
            sg.Text('Width Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Width Max',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Min',size=(11,1),font=('Helvetica',15), text_color='red'), 
            sg.Text('Height Max',size=(12,1),font=('Helvetica',15), text_color='red'),
            sg.Text('PLC',size=(11,1),font=('Helvetica',15), text_color='red'),
            sg.Text('Confidence',size=(11,1),font=('Helvetica',15), text_color='red')],

        ], relief=sg.RELIEF_FLAT)],
        [sg.Frame('',[
            [
                sg.Text(f'{model4.names[i4]}_4',size=(15,1),font=('Helvetica',15), text_color='yellow'), 
                sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key=f'{model4.names[i4]}_4',enable_events=True, disabled=True), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model4.names[i4]}_OK_4',enable_events=True, disabled=True), 
                sg.Input('1',size=(2,1),font=('Helvetica',15),key= f'{model4.names[i4]}_Num_4',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(4,1),font=('Helvetica',15), text_color='red'), 
                sg.Checkbox('',size=(5,5),font=('Helvetica',15),  key=f'{model4.names[i4]}_NG_4',enable_events=True, disabled=True), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model4.names[i4]}_Wn_4',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model4.names[i4]}_Wx_4',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model4.names[i4]}_Hn_4',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('100000',size=(8,1),font=('Helvetica',15),key= f'{model4.names[i4]}_Hx_4',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Input('0',size=(8,1),font=('Helvetica',15),key= f'{model4.names[i4]}_PLC_4',text_color='navy',enable_events=True, disabled=True), 
                sg.Text('',size=(2,1),font=('Helvetica',15), text_color='red'), 
                sg.Slider(range=(1,100),default_value=25,orientation='h',size=(30,20),font=('Helvetica',11), key= f'{model4.names[i4]}_Conf_4',enable_events=True, disabled=True),            
                
            ] for i4 in range(len(model4.names))
        ], relief=sg.RELIEF_FLAT)],
        [sg.Text('  OK',size=(15,1),font=('Helvetica',15), text_color='yellow'),
        sg.Text(' '*230), 
        sg.Input('0',size=(8,1),font=('Helvetica',15),key= 'OK_PLC_4',text_color='navy',enable_events=True)],
        [sg.Text(' ')],
        [sg.Text(' '*250), sg.Button('Save Data', size=(12,1),  font=('Helvetica',12),key='SaveData4',enable_events=True)] 
        ])]
    ]





    layout_saveimg = [
        [sg.Frame('',[
                [sg.Text('Have save folder image OK for camera 1',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_1',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 1', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam1/OK' ,font=('Helvetica',12), key='save_OK_1',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_1',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image OK for camera 2',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_2',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 2', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam2/OK' , font=('Helvetica',12), key='save_OK_2',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_2',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 1',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_1',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 1', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam1/NG' , font=('Helvetica',12), key='save_NG_1',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_1',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 2',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_2',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 2', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam2/NG' , font=('Helvetica',12), key='save_NG_2',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_2',enable_events=True) ],
        ], relief=sg.RELIEF_FLAT),
        sg.Frame('',[
                [sg.Text('Have save folder image OK for camera 3',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_3',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 3', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam3/OK' ,font=('Helvetica',12), key='save_OK_3',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_3',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image OK for camera 4',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=False,font=('Helvetica',15),  key='have_save_OK_4',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image OK for camera 4', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam4/OK' , font=('Helvetica',12), key='save_OK_4',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_OK_4',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 3',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_3',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 3', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam3/NG' , font=('Helvetica',12), key='save_NG_3',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_3',enable_events=True) ],
                [sg.Text('')],
                [sg.Text('Have save folder image NG for camera 4',size=(35,1),font=('Helvetica',15), text_color='yellow'),sg.Checkbox('',size=(5,5),default=True,font=('Helvetica',15),  key='have_save_NG_4',enable_events=True, disabled=True)], 
                [sg.T('Choose folder save image NG for camera 4', font='Any 15', text_color = 'green')],
                [sg.Input(size=(50,1),default_text='C:/Cam4/NG' , font=('Helvetica',12), key='save_NG_4',readonly= True, text_color='navy',enable_events= True),
                sg.FolderBrowse(size=(12,1), font=('Helvetica',10),key='save_folder_NG_4',enable_events=True) ],
        ], relief=sg.RELIEF_FLAT)],
        ]
    layout_terminal = [[sg.Text("Anything printed will display here!")],
                      [sg.Multiline( font=('Helvetica',14), write_only=True, autoscroll=True, auto_refresh=True,reroute_stdout=True, reroute_stderr=True, echo_stdout_stderr=True,expand_x=True,expand_y=True)]
                      ]
    
    layout = [[sg.TabGroup([[  sg.Tab('Main', layout_main),
                               sg.Tab('Option for model 1a', layout_option1a),
                                
                               sg.Tab('Option for model 1b', layout_option1b),
                               sg.Tab('Option for model 2', layout_option2),
                               sg.Tab('Option for model 3', layout_option3),
                               sg.Tab('Option for model 4', layout_option4),
                               sg.Tab('Save Image', layout_saveimg),
                               sg.Tab('Output', layout_terminal)]])
               ]]

    #layout[-1].append(sg.Sizegrip())
    window = sg.Window('HuynhLeVu', layout, location=(0,0),right_click_menu=right_click_menu,resizable=True).Finalize()
    #window.bind('<Configure>',"Configure")
    window.Maximize()

    return window




image_width_display = int(650/1.4)
image_height_display = int(410/1.4)

result_width_display = int(70/1.4)
result_height_display = int(100/1.4)


file_name_img = [("Img(*.jpg,*.png)",("*jpg","*.png"))]


recording1 = False
recording2 = False 

error_cam1 = True
error_cam2 = True

recording3 = False
recording4 = False 

error_cam3 = True
error_cam4 = True

#window['result_cam1'].update(value= 'Wait', text_color='yellow')
#window['result_cam2'].update(value= 'Wait', text_color='yellow')



#moplc
soc = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
connected = False
while connected == False:
    connected = socket_connect('192.168.0.10',8501)
print("connected") 
######################################################end

mypath1 = load_model(1)
model1 = torch.hub.load('./levu','custom', path= mypath1, source='local',force_reload =False)

img1_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result1 = model1(img1_test,416,0.25) 
print('model1 already')


mypath2 = load_model(2)
model2 = torch.hub.load('./levu','custom', path= mypath2, source='local',force_reload =False)

img2_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result2 = model2(img2_test,416,0.25) 

print('model2 already')


mypath3 = load_model(3)
model3 = torch.hub.load('./levu','custom', path= mypath3, source='local',force_reload =False)

img1_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result3 = model3(img1_test,416,0.25) 
print('model3 already')


mypath4 = load_model(4)
model4 = torch.hub.load('./levu','custom', path= mypath4, source='local',force_reload =False)

img2_test = os.path.join(os.getcwd(), 'img/imgtest.jpg')
result4 = model4(img2_test,416,0.25) 
print('model4 already')

choose_model = load_choosemodel()

themes = load_theme()
theme = themes[0]
window = make_window(theme)

window['choose_model'].update(value=choose_model)


try:
    load_all_sql(1,choose_model)
except:
    print(traceback.format_exc())
    window['time_cam1'].update(value= "Error data") 


try:
    load_all_sql(2,choose_model)
except:
    print(traceback.format_exc())
    window['time_cam2'].update(value= "Error data") 

try:
    load_all_sql(3,choose_model)
except:
    print(traceback.format_exc())
    window['time_cam3'].update(value= "Error data") 


try:
    load_all_sql(4,choose_model)
except:
    print(traceback.format_exc())
    window['time_cam4'].update(value= "Error data") 


connect_camera1 = False
connect_camera2 = False
connect_camera3 = False
connect_camera4 = False
connect_total = False

removefile()

#moplc
writedata('DM1900.U',0) 
writedata('DM1902.U',0) 
writedata('DM1904.U',0) 
writedata('DM1906.U',0) 
######################################################end


if connect_camera1 == True and connect_total == True:
    window['result_cam1'].update(value= 'Done', text_color='blue')
if connect_camera2 == True and connect_total == True:
    window['result_cam2'].update(value= 'Done', text_color='blue')
if connect_camera3 == True and connect_total == True:
    window['result_cam3'].update(value= 'Done', text_color='blue')
if connect_camera4 == True and connect_total == True:
    window['result_cam4'].update(value= 'Done', text_color='blue')

#removefile()
try:
    while True:
        event, values = window.read(timeout=20)

        # for i1 in range(len(model1.names)):
        #     #if event == f'{model1.names[i1]}_1':
        #     if values[f'{model1.names[i1]}_1'] == False:
        #         window[f'{model1.names[i1]}_OK_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_Num_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_NG_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_Wn_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_Wx_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_Hn_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_Hx_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_PLC_1'].update(disabled=True)
        #         window[f'{model1.names[i1]}_Conf_1'].update(disabled=True)


        #     elif values[f'{model1.names[i1]}_1'] == True:
        #         window[f'{model1.names[i1]}_OK_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_Num_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_Wn_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_Wx_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_Hn_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_Hx_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_PLC_1'].update(disabled=False)
        #         window[f'{model1.names[i1]}_Conf_1'].update(disabled=False)


        # for i1 in range(len(model1.names)):
        #     if event == f'{model1.names[i1]}_OK_1':
        #         if values[f'{model1.names[i1]}_OK_1'] == True:
        #             window[f'{model1.names[i1]}_NG_1'].update(disabled=True)
        #         else:
        #             window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
        #     if event == f'{model1.names[i1]}_NG_1':
        #         if values[f'{model1.names[i1]}_NG_1'] == True:
        #             window[f'{model1.names[i1]}_OK_1'].update(disabled=True)
        #         else:
        #             window[f'{model1.names[i1]}_OK_1'].update(disabled=False)


        # for i2 in range(len(model2.names)):
        #     #if event == f'{model2.names[i2]}_2':
        #     if values[f'{model2.names[i2]}_2'] == False:
        #         window[f'{model2.names[i2]}_OK_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_Num_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_NG_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_Wn_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_Wx_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_Hn_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_Hx_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_PLC_2'].update(disabled=True)
        #         window[f'{model2.names[i2]}_Conf_2'].update(disabled=True)


        #     elif values[f'{model2.names[i2]}_2'] == True:
        #         window[f'{model2.names[i2]}_OK_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_Num_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_Wn_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_Wx_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_Hn_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_Hx_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_PLC_2'].update(disabled=False)
        #         window[f'{model2.names[i2]}_Conf_2'].update(disabled=False)


        # for i2 in range(len(model2.names)):
        #     if event == f'{model2.names[i2]}_OK_2':
        #         if values[f'{model2.names[i2]}_OK_2'] == True:
        #             window[f'{model2.names[i2]}_NG_2'].update(disabled=True)
        #         else:
        #             window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
        #     if event == f'{model2.names[i2]}_NG_2':
        #         if values[f'{model2.names[i2]}_NG_2'] == True:
        #             window[f'{model2.names[i2]}_OK_2'].update(disabled=True)
        #         else:
        #             window[f'{model2.names[i2]}_OK_2'].update(disabled=False)

        # for i3 in range(len(model3.names)):
        #     #if event == f'{model3.names[i3]}_3':
        #     if values[f'{model3.names[i3]}_3'] == False:
        #         window[f'{model3.names[i3]}_OK_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_Num_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_NG_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_Wn_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_Wx_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_Hn_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_Hx_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_PLC_3'].update(disabled=True)
        #         window[f'{model3.names[i3]}_Conf_3'].update(disabled=True)


        #     elif values[f'{model3.names[i3]}_3'] == True:
        #         window[f'{model3.names[i3]}_OK_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_Num_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_Wn_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_Wx_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_Hn_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_Hx_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_PLC_3'].update(disabled=False)
        #         window[f'{model3.names[i3]}_Conf_3'].update(disabled=False)


        # for i3 in range(len(model3.names)):
        #     if event == f'{model3.names[i3]}_OK_3':
        #         if values[f'{model3.names[i3]}_OK_3'] == True:
        #             window[f'{model3.names[i3]}_NG_3'].update(disabled=True)
        #         else:
        #             window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
        #     if event == f'{model3.names[i3]}_NG_3':
        #         if values[f'{model3.names[i3]}_NG_3'] == True:
        #             window[f'{model3.names[i3]}_OK_3'].update(disabled=True)
        #         else:
        #             window[f'{model3.names[i3]}_OK_3'].update(disabled=False)



        # for i4 in range(len(model4.names)):
        #     #if event == f'{model4.names[i4]}_4':
        #     if values[f'{model4.names[i4]}_4'] == False:
        #         window[f'{model4.names[i4]}_OK_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_Num_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_NG_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_Wn_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_Wx_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_Hn_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_Hx_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_PLC_4'].update(disabled=True)
        #         window[f'{model4.names[i4]}_Conf_4'].update(disabled=True)


        #     elif values[f'{model4.names[i4]}_4'] == True:
        #         window[f'{model4.names[i4]}_OK_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_Num_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_NG_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_Wn_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_Wx_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_Hn_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_Hx_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_PLC_4'].update(disabled=False)
        #         window[f'{model4.names[i4]}_Conf_4'].update(disabled=False)


        # for i4 in range(len(model4.names)):
        #     if event == f'{model4.names[i4]}_OK_4':
        #         if values[f'{model4.names[i4]}_OK_4'] == True:
        #             window[f'{model4.names[i4]}_NG_4'].update(disabled=True)
        #         else:
        #             window[f'{model4.names[i4]}_NG_4'].update(disabled=False)
        #     if event == f'{model4.names[i4]}_NG_4':
        #         if values[f'{model4.names[i4]}_NG_4'] == True:
        #             window[f'{model4.names[i4]}_OK_4'].update(disabled=True)
        #         else:
        #             window[f'{model4.names[i4]}_OK_4'].update(disabled=False)

        if event =='Exit' or event == sg.WINDOW_CLOSED :
            break

        # if event == 'Configure':
        #     if window.TKroot.state() == 'zoomed':
        #         #print(window['image1'].get_size()[0])
        #         image_width_display = window['image1'].get_size()[0]
        #         image_height_display = window['image1'].get_size()[1]
        #         result_width_display = image_width_display - 190
        #         result_height_display = 100 


        if event =='Administrator':
            login_password = 'vu123'  # helloworld
            password = sg.popup_get_text(
                'Enter Password: ', password_char='*') 
            if password == login_password:
                sg.popup_ok('Login Successed!!! ',text_color='green', font=('Helvetica',14))  

                window['conf_thres2'].update(disabled= False)
                window['conf_thres1'].update(disabled= False)

                window['file_browse2'].update(disabled= False,button_color='turquoise')
                window['file_browse1'].update(disabled= False,button_color='turquoise')

                window['SaveData1'].update(disabled= False,button_color='turquoise')
                window['SaveData2'].update(disabled= False,button_color='turquoise')

                window['Webcam1'].update(disabled= False,button_color='turquoise')
                window['Webcam2'].update(disabled= False,button_color='turquoise')
                window['Stop1'].update(disabled= False,button_color='turquoise')
                window['Stop2'].update(disabled= False,button_color='turquoise')
                window['Pic1'].update(disabled= False,button_color='turquoise')
                window['Pic2'].update(disabled= False,button_color='turquoise')
                window['Snap1'].update(disabled= False,button_color='turquoise')
                window['Snap2'].update(disabled= False,button_color='turquoise')
                window['Change1'].update(button_color='turquoise')
                window['Change2'].update(button_color='turquoise')
                window['Change_1'].update(button_color='turquoise')
                window['Change_2'].update(button_color='turquoise')
                window['Detect1'].update(button_color='turquoise')
                window['Detect2'].update(button_color='turquoise')


                window['have_save_OK_1'].update(disabled=False)
                window['have_save_NG_1'].update(disabled=False)
                window['have_save_OK_2'].update(disabled=False)
                window['have_save_NG_2'].update(disabled=False)

                window['save_OK_1'].update(disabled=False)
                window['save_NG_1'].update(disabled=False)
                window['save_OK_2'].update(disabled=False)
                window['save_NG_2'].update(disabled=False)

                window['save_folder_OK_1'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_1'].update(disabled= False,button_color='turquoise')
                window['save_folder_OK_2'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_2'].update(disabled= False,button_color='turquoise')


                for i1 in range(len(model1.names)):
                    window[f'{model1.names[i1]}_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_OK_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Num_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_NG_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Wn_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Wx_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Hn_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Hx_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_PLC_1'].update(disabled=False)
                    window[f'{model1.names[i1]}_Conf_1'].update(disabled=False)


                for i2 in range(len(model2.names)):
                    window[f'{model2.names[i2]}_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_OK_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Num_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_NG_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Wn_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Wx_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Hn_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Hx_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_PLC_2'].update(disabled=False)
                    window[f'{model2.names[i2]}_Conf_2'].update(disabled=False)



                window['conf_thres4'].update(disabled= False)
                window['conf_thres3'].update(disabled= False)

                window['file_browse4'].update(disabled= False,button_color='turquoise')
                window['file_browse3'].update(disabled= False,button_color='turquoise')

                window['SaveData3'].update(disabled= False,button_color='turquoise')
                window['SaveData4'].update(disabled= False,button_color='turquoise')

                window['Webcam3'].update(disabled= False,button_color='turquoise')
                window['Webcam4'].update(disabled= False,button_color='turquoise')
                window['Stop3'].update(disabled= False,button_color='turquoise')
                window['Stop4'].update(disabled= False,button_color='turquoise')
                window['Pic3'].update(disabled= False,button_color='turquoise')
                window['Pic4'].update(disabled= False,button_color='turquoise')
                window['Snap3'].update(disabled= False,button_color='turquoise')
                window['Snap4'].update(disabled= False,button_color='turquoise')
                window['Change3'].update(button_color='turquoise')
                window['Change4'].update(button_color='turquoise')
                window['Change_3'].update(button_color='turquoise')
                window['Change_4'].update(button_color='turquoise')
                window['Detect3'].update(button_color='turquoise')
                window['Detect4'].update(button_color='turquoise')


                window['have_save_OK_3'].update(disabled=False)
                window['have_save_NG_3'].update(disabled=False)
                window['have_save_OK_4'].update(disabled=False)
                window['have_save_NG_4'].update(disabled=False)

                window['save_OK_3'].update(disabled=False)
                window['save_NG_3'].update(disabled=False)
                window['save_OK_4'].update(disabled=False)
                window['save_NG_4'].update(disabled=False)

                window['save_folder_OK_3'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_3'].update(disabled= False,button_color='turquoise')
                window['save_folder_OK_4'].update(disabled= False,button_color='turquoise')
                window['save_folder_NG_4'].update(disabled= False,button_color='turquoise')


                for i3 in range(len(model3.names)):
                    window[f'{model3.names[i3]}_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_OK_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Num_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_NG_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Wn_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Wx_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Hn_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Hx_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_PLC_3'].update(disabled=False)
                    window[f'{model3.names[i3]}_Conf_3'].update(disabled=False)


                for i4 in range(len(model4.names)):
                    window[f'{model4.names[i4]}_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_OK_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_Num_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_NG_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_Wn_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_Wx_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_Hn_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_Hx_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_PLC_4'].update(disabled=False)
                    window[f'{model4.names[i4]}_Conf_4'].update(disabled=False)

                    


            else:
                sg.popup_cancel('Wrong Password!!!',text_color='red', font=('Helvetica',14))


        if event == 'Change Theme':
            layout_theme = [
                [sg.Listbox(values= sg.theme_list(), size = (30,20),auto_size_text=18,default_values='Dark',key='theme', enable_events=True)],
                [
                    [sg.Button('Apply'),
                    sg.Button('Cancel')]
                ]
            ] 
            window_theme = sg.Window('Change Theme', layout_theme, location=(50,50),keep_on_top=True).Finalize()
            window_theme.set_min_size((300,400))   

            while True:
                event_theme, values_theme = window_theme.read(timeout=20)
                if event_theme == sg.WIN_CLOSED:
                    break

                if event_theme == 'Apply':
                    theme_choose = values_theme['theme'][0]
                    if theme_choose == 'Default':
                        continue
                    window.close()
                    window = make_window(theme_choose)
                    save_theme(theme_choose)
                    #print(theme_choose)
                if event_theme == 'Cancel':
                    answer = sg.popup_yes_no('Do you want to exit?')
                    if answer == 'Yes':
                        break
                    if answer == 'No':
                        continue
            window_theme.close()



        if event == 'file_browse1': 
            window['file_weights1'].update(value=values['file_browse1'])
            if values['file_browse1']:
                window['Change1'].update(disabled=False)
                window['Change_1'].update(disabled=False)



        if event == 'file_browse2':
            window['file_weights2'].update(value=values['file_browse2'])
            if values['file_browse2']:
                window['Change2'].update(disabled=False)
                window['Change_2'].update(disabled=False)


        if event == 'file_browse3': 
            window['file_weights3'].update(value=values['file_browse3'])
            if values['file_browse3']:
                window['Change3'].update(disabled=False)
                window['Change_3'].update(disabled=False)



        if event == 'file_browse4':
            window['file_weights4'].update(value=values['file_browse4'])
            if values['file_browse4']:
                window['Change4'].update(disabled=False)
                window['Change_4'].update(disabled=False)



        if event == 'choose_model':
            mychoose = values['choose_model']
            weight1 = ''
            conf_thres1 = 1
            weight2 = ''
            conf_thres2 = 1

            OK_Cam1 = False
            OK_Cam2 = False
            NG_Cam1 = True
            NG_Cam2 = True
            Folder_OK_Cam1 = 'C:/Cam1/OK'
            Folder_OK_Cam2 = 'C:/Cam2/OK'
            Folder_NG_Cam1 = 'C:/Cam1/NG'
            Folder_NG_Cam2 = 'C:/Cam2/NG'

            weight3 = ''
            conf_thres3 = 1
            weight4 = ''
            conf_thres4 = 1

            OK_Cam3 = False
            OK_Cam4 = False
            NG_Cam3 = True
            NG_Cam4 = True
            Folder_OK_Cam3 = 'C:/Cam3/OK'
            Folder_OK_Cam4 = 'C:/Cam4/OK'
            Folder_NG_Cam3 = 'C:/Cam3/NG'
            Folder_NG_Cam4 = 'C:/Cam4/NG'

            conn = sqlite3.connect('modeldb_4_PLC_conf.db')
            cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,OK_Cam4,NG_Cam1,NG_Cam2,NG_Cam3,NG_Cam4,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_OK_Cam4,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Folder_NG_Cam4,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax from MYMODEL")
            for row in cursor:
                if row[0] == values['choose_model']:
 
                    mychoose = values['choose_model']
                    row1_a, row1_b = row[1].strip().split('_')
                    if row1_a == '1' and row1_b == '0':
                        weight1 = row[2]
                        conf_thres1 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
                        OK_Cam4 = str2bool(row[7])
                        NG_Cam1 = str2bool(row[8])
                        NG_Cam2 = str2bool(row[9])
                        NG_Cam3 = str2bool(row[10])
                        NG_Cam4 = str2bool(row[11])
                        Folder_OK_Cam1 = row[12]
                        Folder_OK_Cam2 = row[13]
                        Folder_OK_Cam3 = row[14]
                        Folder_OK_Cam4 = row[15]
                        Folder_NG_Cam1 = row[16]
                        Folder_NG_Cam2 = row[17]
                        Folder_NG_Cam3 = row[18]
                        Folder_NG_Cam4 = row[19]
                        model1 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

                    if row1_a == '2' and row1_b == '0':
                        weight2 = row[2]
                        conf_thres2 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
                        OK_Cam4 = str2bool(row[7])
                        NG_Cam1 = str2bool(row[8])
                        NG_Cam2 = str2bool(row[9])
                        NG_Cam3 = str2bool(row[10])
                        NG_Cam4 = str2bool(row[11])
                        Folder_OK_Cam1 = row[12]
                        Folder_OK_Cam2 = row[13]
                        Folder_OK_Cam3 = row[14]
                        Folder_OK_Cam4 = row[15]
                        Folder_NG_Cam1 = row[16]
                        Folder_NG_Cam2 = row[17]
                        Folder_NG_Cam3 = row[18]
                        Folder_NG_Cam4 = row[19]
                        model2 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

                    if row1_a == '3' and row1_b == '0':
                        weight3 = row[2]
                        conf_thres3 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
                        OK_Cam4 = str2bool(row[7])
                        NG_Cam1 = str2bool(row[8])
                        NG_Cam2 = str2bool(row[9])
                        NG_Cam3 = str2bool(row[10])
                        NG_Cam4 = str2bool(row[11])
                        Folder_OK_Cam1 = row[12]
                        Folder_OK_Cam2 = row[13]
                        Folder_OK_Cam3 = row[14]
                        Folder_OK_Cam4 = row[15]
                        Folder_NG_Cam1 = row[16]
                        Folder_NG_Cam2 = row[17]
                        Folder_NG_Cam3 = row[18]
                        Folder_NG_Cam4 = row[19]
                        model3 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)

                    if row1_a == '4' and row1_b == '0':
                        weight4 = row[2]
                        #window['conf_thres2'].update(value=row[3])
                        conf_thres4 = row[3]
                        OK_Cam1 = str2bool(row[4])
                        OK_Cam2 = str2bool(row[5])
                        OK_Cam3 = str2bool(row[6])
                        OK_Cam4 = str2bool(row[7])
                        NG_Cam1 = str2bool(row[8])
                        NG_Cam2 = str2bool(row[9])
                        NG_Cam3 = str2bool(row[10])
                        NG_Cam4 = str2bool(row[11])
                        Folder_OK_Cam1 = row[12]
                        Folder_OK_Cam2 = row[13]
                        Folder_OK_Cam3 = row[14]
                        Folder_OK_Cam4 = row[15]
                        Folder_NG_Cam1 = row[16]
                        Folder_NG_Cam2 = row[17]
                        Folder_NG_Cam3 = row[18]
                        Folder_NG_Cam4 = row[19]
                        model4 = torch.hub.load('./levu','custom', path= row[2], source='local',force_reload =False)
        
            window.close() 
            window = make_window(theme)

            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)
            window['choose_model'].update(value=mychoose)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)

            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)
            window['file_weights4'].update(value=weight4)
            window['conf_thres4'].update(value=conf_thres4)
            window['choose_model'].update(value=mychoose)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_OK_4'].update(value=OK_Cam4)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['have_save_NG_4'].update(value=NG_Cam4)

            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_OK_4'].update(value=Folder_OK_Cam4)
            window['save_NG_3'].update(value=Folder_NG_Cam3)
            window['save_NG_4'].update(value=Folder_NG_Cam4)


            cursor = conn.execute("SELECT ChooseModel,Camera,Weights,Confidence,OK_Cam1,OK_Cam2,OK_Cam3,OK_Cam4,NG_Cam1,NG_Cam2,NG_Cam3,NG_Cam4,Folder_OK_Cam1,Folder_OK_Cam2,Folder_OK_Cam3,Folder_OK_Cam4,Folder_NG_Cam1,Folder_NG_Cam2,Folder_NG_Cam3,Folder_NG_Cam4,Joined,Ok,Num,NG,WidthMin,WidthMax,HeightMin,HeightMax,PLC_NG,PLC_OK,Conf from MYMODEL")
            for row in cursor:
                if row[0] == values['choose_model']:
                    row1_a, row1_b = row[1].strip().split('_')
                    if row1_a == '1':
                        for item in range(len(model1.names)):
                            if int(row1_b) == item:
                                window[f'{model1.names[item]}_1'].update(value=str2bool(row[20]))
                                window[f'{model1.names[item]}_OK_1'].update(value=str2bool(row[21]))
                                window[f'{model1.names[item]}_Num_1'].update(value=str(row[22]))
                                window[f'{model1.names[item]}_NG_1'].update(value=str2bool(row[23]))
                                window[f'{model1.names[item]}_Wn_1'].update(value=str(row[24]))
                                window[f'{model1.names[item]}_Wx_1'].update(value=str(row[25]))
                                window[f'{model1.names[item]}_Hn_1'].update(value=str(row[26]))
                                window[f'{model1.names[item]}_Hx_1'].update(value=str(row[27]))
                                window[f'{model1.names[item]}_PLC_1'].update(value=str(row[28]))
                                window['OK_PLC_1'].update(value=str(row[29]))
                                window[f'{model1.names[item]}_Conf_1'].update(value=str(row[30]))
                    if row1_a == '2':
                        for item in range(len(model2.names)):
                            if int(row1_b) == item:
                                window[f'{model2.names[item]}_2'].update(value=str2bool(row[20]))
                                window[f'{model2.names[item]}_OK_2'].update(value=str2bool(row[21]))
                                window[f'{model2.names[item]}_Num_2'].update(value=str(row[22]))
                                window[f'{model2.names[item]}_NG_2'].update(value=str2bool(row[23]))
                                window[f'{model2.names[item]}_Wn_2'].update(value=str(row[24]))
                                window[f'{model2.names[item]}_Wx_2'].update(value=str(row[25]))
                                window[f'{model2.names[item]}_Hn_2'].update(value=str(row[26]))
                                window[f'{model2.names[item]}_Hx_2'].update(value=str(row[27]))
                                window[f'{model2.names[item]}_PLC_2'].update(value=str(row[28]))
                                window['OK_PLC_2'].update(value=str(row[29]))
                                window[f'{model2.names[item]}_Conf_2'].update(value=str(row[30]))
                    if row1_a == '3':
                        for item in range(len(model3.names)):
                            if int(row1_b) == item:
                                window[f'{model3.names[item]}_3'].update(value=str2bool(row[20]))
                                window[f'{model3.names[item]}_OK_3'].update(value=str2bool(row[21]))
                                window[f'{model3.names[item]}_Num_3'].update(value=str(row[22]))
                                window[f'{model3.names[item]}_NG_3'].update(value=str2bool(row[23]))
                                window[f'{model3.names[item]}_Wn_3'].update(value=str(row[24]))
                                window[f'{model3.names[item]}_Wx_3'].update(value=str(row[25]))
                                window[f'{model3.names[item]}_Hn_3'].update(value=str(row[26]))
                                window[f'{model3.names[item]}_Hx_3'].update(value=str(row[27]))
                                window[f'{model3.names[item]}_PLC_3'].update(value=str(row[28]))
                                window['OK_PLC_3'].update(value=str(row[29]))
                                window[f'{model3.names[item]}_Conf_3'].update(value=str(row[30]))

                    if row1_a == '4':
                        for item in range(len(model4.names)):
                            if int(row1_b) == item:
                                window[f'{model4.names[item]}_4'].update(value=str2bool(row[20]))
                                window[f'{model4.names[item]}_OK_4'].update(value=str2bool(row[21]))
                                window[f'{model4.names[item]}_Num_4'].update(value=str(row[22]))
                                window[f'{model4.names[item]}_NG_4'].update(value=str2bool(row[23]))
                                window[f'{model4.names[item]}_Wn_4'].update(value=str(row[24]))
                                window[f'{model4.names[item]}_Wx_4'].update(value=str(row[25]))
                                window[f'{model4.names[item]}_Hn_4'].update(value=str(row[26]))
                                window[f'{model4.names[item]}_Hx_4'].update(value=str(row[27]))
                                window[f'{model4.names[item]}_PLC_4'].update(value=str(row[28]))
                                window['OK_PLC_4'].update(value=str(row[29]))
                                window[f'{model4.names[item]}_Conf_4'].update(value=str(row[30]))

            conn.close()


        if event == 'SaveData1':

            save_all_sql(model1,1,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(1,values['file_weights1'])
            sg.popup('Saved param model 1 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)


        if event == 'SaveData2':
            save_all_sql(model2,2,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(2,values['file_weights2'])
            sg.popup('Saved param model 2 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)


        if event == 'SaveData3':

            save_all_sql(model3,3,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(3,values['file_weights3'])
            sg.popup('Saved param model 3 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)


        if event == 'SaveData4':
            save_all_sql(model4,4,str(values['choose_model']))
            save_choosemodel(values['choose_model'])
            save_model(4,values['file_weights4'])
            sg.popup('Saved param model 4 successed',font=('Helvetica',15), text_color='green',keep_on_top= True)

        #moplc
        program_camera1_FH(model=model1,size= 608,conf= values['conf_thres1']/100)
        program_camera2_FH(model=model2,size= 416,conf= values['conf_thres2']/100)
        program_camera3_FH(model=model3,size= 416,conf= values['conf_thres3']/100)
        program_camera4_FH(model=model4,size= 416,conf= values['conf_thres4']/100)

        if readdata('DM1920') == 1:
            removefile()
            writedata('DM1920.U',0) 







        try:
            for var_plc in range(32):
                if readdata(f'DM252{var_plc}') == 1:
                    excel_pp(var_plc)
                    writedata(f'DM252{var_plc}',0) 
        except:
            pass




        try:
            #HomNay
            if readdata('DM6050') == 1:
                excel_sang()

            # #Dem
            if readdata('DM6050') == 2:
                excel_dem()


            if readdata('DM1100') == 1:
                today = date.today()
                d1 = today.strftime("%d/%m/%Y")
                now = datetime.datetime.now()
                t1 = now.strftime("%H:%M:%S")

                mydate = today.strftime("%Y_%m_%d")
            
                hour = int(now.strftime("%H"))
        
                if 7<= hour <=18:
                    if not os.path.isfile(f"excel/{mydate}_Ngay.xlsx"):
                        excel_sang()
                    wb = openpyxl.load_workbook(f"excel/{mydate}_Ngay.xlsx")

                if 19 <= hour <= 23:
                    if not os.path.isfile(f"excel/{mydate}_Dem.xlsx"):
                        excel_dem()
                    wb = openpyxl.load_workbook(f"excel/{mydate}_Dem.xlsx")

                if 0 <= hour <= 6:
                    Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
                    Previous_Date = Previous_Date.strftime("%Y_%m_%d")
                    if not os.path.isfile(f"excel/{Previous_Date}_Dem.xlsx"):
                        excel_dem()

                    wb = openpyxl.load_workbook(f"excel/{Previous_Date}_Dem.xlsx")


                ws = wb.active
                col1 = d1
                col2 = t1
                col3 = int(readdata('DM110'))
                col4 = int(readdata('DM6000'))
                col5 = int(readdata('DM6004'))
                col6 = int(readdata('DM6008'))
                col7 = int(readdata('DM6012'))
                col8 = int(readdata('DM6016'))
                col9 = int(readdata('DM6020'))
                col10 = int(readdata('DM6024'))
                col11 = int(readdata('DM6032'))
                col12 = int(readdata('DM6028'))
                col13 = int(readdata('DM6042'))
                ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12,col13,])

                for row in range(ws.max_row, ws.max_row+1):
                    for col in range(1, ws.max_column+1):
                        #print(col[row].value)
                        d = ws.cell(row = row, column = col)
                        #currentCell = ws.cell(col[row])
                        d.alignment = Alignment(horizontal='center')
                        #d.style.alignment.horizontal = 'center'
                        d.font = Font(name= 'Calibri', size=12)

                if 7<= hour <=18:
                    
                    wb.save(f"excel/{mydate}_Ngay.xlsx")
                    try:
                        shutil.copy(f"excel/{mydate}_Ngay.xlsx", f"C:/Users/Administrator/Desktop/excel/{mydate}_Ngay.xlsx")
                    except:
                        pass
                    
                    try:
                        wb1 = openpyxl.load_workbook(f"C:/Users/Administrator/Desktop/excel/{mydate}_Ngay.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 14):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/Users/Administrator/Desktop/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 14):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 14):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/Users/Administrator/Desktop/excel/Now.xlsx")



                    except:
                        pass
                if 19 <= hour <= 23:
                    wb.save(f"excel/{mydate}_Dem.xlsx")
                    try:
                        shutil.copy(f"excel/{mydate}_Dem.xlsx", f"C:/Users/Administrator/Desktop/excel/{mydate}_Dem.xlsx")
                    except:
                        pass

                    try:
                        wb1 = openpyxl.load_workbook(f"C:/Users/Administrator/Desktop/excel/{mydate}_Dem.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 14):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/Users/Administrator/Desktop/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 14):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 14):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/Users/Administrator/Desktop/excel/Now.xlsx")



                    except:
                        pass


                if 0 <= hour <= 6:
                    Previous_Date = datetime.datetime.today() - datetime.timedelta(days=1)
                    Previous_Date = Previous_Date.strftime("%Y_%m_%d")
                    wb.save(f"excel/{Previous_Date}_Dem.xlsx")
                    try:
                        shutil.copy(f"excel/{Previous_Date}_Dem.xlsx", f"C:/Users/Administrator/Desktop/excel/{Previous_Date}_Dem.xlsx")
                    except:
                        pass

                    try:
                        wb1 = openpyxl.load_workbook(f"C:/Users/Administrator/Desktop/excel/{Previous_Date}_Dem.xlsx")

                        ws1 = wb1.active
                        all_cell = []


                        for row in range(1, ws1.max_row+1):
                            all_col = []    
                            for col in range(1, 14):
                                cell_obj = ws1.cell(row = row, column = col)
                                all_col.append(cell_obj.value)
                            all_cell.append(all_col)




                        wb2 = openpyxl.load_workbook("C:/Users/Administrator/Desktop/excel/Now.xlsx")
                        ws2 = wb2.active

                        index_row = ws2.max_row


                        if ws1.max_row > ws2.max_row:
                            for row in range(4,ws1.max_row+1):
                                for col in range(1, 14):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                    
                        else:
                            for row in range(4,ws2.max_row+1):
                                for col in range(1, 14):

                                    if len(all_cell) >= row:
                                        ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)
                                    else:
                                        ws2.cell(row = row, column = col).value = None
                                        d = ws2.cell(row = row, column = col)
                                        d.alignment = Alignment(horizontal='center')
                                        d.font = Font(name= 'Calibri', size=12)   

                        wb2.save("C:/Users/Administrator/Desktop/excel/Now.xlsx")



                    except:
                        pass


                writedata('DM1100.U',0) 
            

            if readdata('DM6054') == 1:
                today = date.today()
                d1 = today.strftime("%d/%m/%Y")
                now = datetime.datetime.now()
                t1 = now.strftime("%H:%M:%S")

                wb = openpyxl.load_workbook("excel/All.xlsx")
                ws = wb.active
                col1 = d1
                col2 = t1
                col3 = int(readdata('DM110'))
                col4 = int(readdata('DM6000'))
                col5 = int(readdata('DM6004'))
                col6 = int(readdata('DM6008'))
                col7 = int(readdata('DM6012'))
                col8 = int(readdata('DM6016'))
                col9 = int(readdata('DM6020'))
                col10 = int(readdata('DM6024'))
                col11 = int(readdata('DM6032'))
                col12 = int(readdata('DM6028'))
                col13 = int(readdata('DM6042'))
                ws.append([col1, col2,col3, col4,col5, col6,col7, col8,col9, col10,col11, col12,col13,])

                for row in range(ws.max_row, ws.max_row+1):
                    for col in range(1, ws.max_column+1):
                        #print(col[row].value)
                        d = ws.cell(row = row, column = col)
                        #currentCell = ws.cell(col[row])
                        d.alignment = Alignment(horizontal='center')
                        #d.style.alignment.horizontal = 'center'
                        d.font = Font(name= 'Calibri', size=12)


                wb.save("excel/All.xlsx")
                try:
                    shutil.copy("excel/All.xlsx", "C:/Users/Administrator/Desktop/excel/A.xlsx")



                    wb1 = openpyxl.load_workbook("C:/Users/Administrator/Desktop/excel/A.xlsx")

                    ws1 = wb1.active
                    all_cell = []
                    all_cell_date = []
                    all_cell_time = []

                    for row in range(1, ws1.max_row+1):
                        all_col = []    
                        for col in range(1, 14):
                            cell_obj = ws1.cell(row = row, column = col)
                            all_col.append(cell_obj.value)
                        all_cell.append(all_col)
                        all_cell_date.append(all_col[0])
                        all_cell_time.append(all_col[1])



                    wb2 = openpyxl.load_workbook("C:/Users/Administrator/Desktop/excel/All.xlsx")
                    ws2 = wb2.active

                    index_row = ws2.max_row
                    for row in range(4, ws2.max_row+1):
                        if ws2.cell(row=row,column=1).value == None:
                            index_row = row
                            break

                    for row in range(index_row,ws1.max_row+1):
                        for col in range(1, 14):
                            ws2.cell(row = row, column = col).value = all_cell[row-1][col-1]
                            d = ws2.cell(row = row, column = col)
                            d.alignment = Alignment(horizontal='center')
                            d.font = Font(name= 'Calibri', size=12)
                    wb2.save("C:/Users/Administrator/Desktop/excel/All.xlsx")


                except:
                    pass
                
                writedata('DM6054.U',0) 
            
        except:
            pass

        ######################################################end
        







        #task_camera1(model=model1,size= 416,conf= values['conf_thres1']/100)
        #task_camera2(model=model2,size= 416,conf= values['conf_thres2']/100)

        #test_camera1(model=model1,size= 416,conf= values['conf_thres1']/100)
        #test_camera2()

        #task1(model1,size= 416,conf= values['conf_thres1']/100)
        #task2(model2,size= 416,conf= values['conf_thres2']/100) 

        #task1(model,size,conf)
        #task2(model,size,conf) 


        ### threading

        # task_thread1 = threading.Thread(target=program_camera1_FH_test, args=(model1, 416, values['conf_thres1']/100,))
        # task_thread2 = threading.Thread(target=program_camera2_FH_test, args=(model2, 416, values['conf_thres2']/100,))
        # task_thread3 = threading.Thread(target=program_camera3_FH_test, args=(model3, 416, values['conf_thres3']/100,))
        # task_thread4 = threading.Thread(target=program_camera4_FH_test, args=(model4, 416, values['conf_thres4']/100,))

        # task_thread1.start()
        # task_thread2.start()
        # task_thread3.start()
        # task_thread4.start()


        #task_thread1 = threading.Thread(target=abc1)
        #task_thread2 = threading.Thread(target=abc2)

        #task_thread1.setDaemon(True)
        #task_thread2.setDaemon(True)

        #task_thread1.join()
        #task_thread2.join()


        # task_thread1 = multiprocessing.Process(target=program_camera1_FH_test, args=(model1, 416, values['conf_thres1']/100,))
        # task_thread2 = multiprocessing.Process(target=program_camera2_FH_test, args=(model2, 416, values['conf_thres2']/100,))
        # task_thread3 = multiprocessing.Process(target=program_camera3_FH_test, args=(model3, 416, values['conf_thres3']/100,))
        # task_thread4 = multiprocessing.Process(target=program_camera4_FH_test, args=(model4, 416, values['conf_thres4']/100,))

        # task_thread1.start()
        # task_thread2.start()
        # task_thread3.start()
        # task_thread4.start()

        # task_thread1.join()
        # task_thread2.join()
        # task_thread3.join()
        # task_thread4.join()




        if event == 'check_model1' and values['check_model1'] == True:
            check_ok=0
            directory1 = 'C:/Check1/'
            if os.listdir(directory1) == []:
                print('folder 1 empty')
            else:
                print('received folder 1')

                for path1 in glob.glob('C:/Check1/*'):
                    name = path1[9:]

                    img1_orgin = cv2.imread(path1)


                    # img1_orgin = cv2.resize(img1_orgin,(640,480))  

                    img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)     


                    result1 = model1(img1_orgin,size= 416,conf = values['conf_thres1']/100)

                    table1 = result1.pandas().xyxy[0]

                    area_remove1 = []

                    myresult1 =0 

                    for item in range(len(table1.index)):
                        width1 = table1['xmax'][item] - table1['xmin'][item]
                        height1 = table1['ymax'][item] - table1['ymin'][item]
                        conf1 = table1['confidence'][item] * 100

                        #area1 = width1*height1
                        label_name = table1['name'][item]
                        for i1 in range(len(model1.names)):
                            if values[f'{model1.names[i1]}_1'] == True:
                                #if values[f'{model1.names[i1]}_WH'] == True:
                                if label_name == model1.names[i1]:
                                    if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)
                                    elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                        table1.drop(item, axis=0, inplace=True)
                                        area_remove1.append(item)       
                        if values[f'{model1.names[i1]}_1'] == False:
                            if label_name == model1.names[i1]:
                                table1.drop(item, axis=0, inplace=True)
                                area_remove1.append(item)


                    names1 = list(table1['name'])

                    show1 = np.squeeze(result1.render(area_remove1))
                    show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)

                    #ta = time.time()
                    for i1 in range(len(model1.names)):
                        #register_ng = (3002 + i1*2).to_bytes(2, byteorder='big') + b'\x00'

                        if values[f'{model1.names[i1]}_OK_1'] == True:
                            len_name1 = 0
                            for name1 in names1:
                                if name1 == model1.names[i1]:
                                    len_name1 +=1
                            if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                print('NG')
                                #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')
                                myresult1 = 1
                                

                        elif values[f'{model1.names[i1]}_NG_1'] == True:
                            if model1.names[i1] in names1:
                                print('NG')
                                #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')    
                                myresult1 = 1         
                                    

                    if myresult1 == 0:
                        print('OK')
                        check_ok = 1
                        #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
                        cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
                        window['result_cam1'].update(value= 'OK', text_color='green')

                    imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                    window['image1'].update(data= imgbytes1)

                    if check_ok == 1:
                        break


        if event == 'check_model2' and values['check_model2'] == True:
            check_ok=0
            directory2 = 'C:/Check2/'
            if os.listdir(directory2) == []:
                print('folder 2 empty')
            else:
                print('received folder 2')

                for path2 in glob.glob('C:/Check2/*'):
                    name = path2[9:]

                    img2_orgin = cv2.imread(path2)
                    # img2_orgin = cv2.resize(img2_orgin,(640,480))  

                    img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB) 

                    result2 = model2(img2_orgin,size= 416,conf = values['conf_thres2']/100)

                    table2 = result2.pandas().xyxy[0]

                    area_remove2 = []

                    myresult2 =0 

                    for item in range(len(table2.index)):
                        width2 = table2['xmax'][item] - table2['xmin'][item]
                        height2 = table2['ymax'][item] - table2['ymin'][item]
                        conf2 = table2['confidence'][item] * 100

                        #area2 = width2*height2
                        label_name = table2['name'][item]
                        for i2 in range(len(model2.names)):
                            if values[f'{model2.names[i2]}_2'] == True:
                                #if values[f'{model2.names[i2]}_WH'] == True:
                                if label_name == model2.names[i2]:
                                    if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)
                                    elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                        table2.drop(item, axis=0, inplace=True)
                                        area_remove2.append(item)  
                        if values[f'{model2.names[i2]}_2'] == False:
                            if label_name == model2.names[i2]:
                                table2.drop(item, axis=0, inplace=True)
                                area_remove2.append(item)

                    names2 = list(table2['name'])

                    show2 = np.squeeze(result2.render(area_remove2))
                    show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            
                    #ta = time.time()
                    for i2 in range(len(model2.names)):
                        #register_ng = (4002 + i1*2).to_bytes(2, byteorder='big') + b'\x00'
                        if values[f'{model2.names[i2]}_OK_2'] == True:
                            len_name2 = 0
                            for name2 in names2:
                                if name2 == model2.names[i2]:
                                    len_name2 +=1
                            if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                                print('NG')
                                #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam2'].update(value= 'NG', text_color='red')
                                myresult2 = 1
                                break

                        if values[f'{model2.names[i2]}_NG_2'] == True:
                            if model2.names[i2] in names2:
                                print('NG')
                                #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x01',1)
                                cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam2'].update(value= 'NG', text_color='red')    
                                myresult2 = 1      
                                break    

                    if myresult2 == 0:
                        print('OK')
                        check_ok = 1
                        #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(4000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x01',1)
                        cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
                        window['result_cam2'].update(value= 'OK', text_color='green')

                    imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                    window['image2'].update(data= imgbytes2)

                    if check_ok == 1:
                        break




        if event == 'check_model3' and values['check_model3'] == True:
            check_ok=0
            directory3 = 'C:/Check3/'
            if os.listdir(directory3) == []:
                print('folder 3 empty')
            else:
                print('received folder 3')

                for path3 in glob.glob('C:/Check3/*'):
                    name = path3[9:]

                    img3_orgin = cv2.imread(path3)


                    # img3_orgin = cv2.resize(img3_orgin,(640,480))  

                    img3_orgin = cv2.cvtColor(img3_orgin, cv2.COLOR_BGR2RGB)     


                    result3 = model3(img3_orgin,size= 416,conf = values['conf_thres3']/100)

                    table3 = result3.pandas().xyxy[0]

                    area_remove3 = []

                    myresult3 =0 

                    for item in range(len(table3.index)):
                        width3 = table3['xmax'][item] - table3['xmin'][item]
                        height3 = table3['ymax'][item] - table3['ymin'][item]
                        conf3 = table3['confidence'][item] * 100

                        #area1 = width3*height3
                        label_name = table3['name'][item]
                        for i3 in range(len(model3.names)):
                            if values[f'{model3.names[i3]}_3'] == True:
                                #if values[f'{model3.names[i3]}_WH'] == True:
                                if label_name == model3.names[i3]:
                                    if width3 < int(values[f'{model3.names[i3]}_Wn_3']): 
                                        table3.drop(item, axis=0, inplace=True)
                                        area_remove3.append(item)
                                    elif width3 > int(values[f'{model3.names[i3]}_Wx_3']): 
                                        table3.drop(item, axis=0, inplace=True)
                                        area_remove3.append(item)
                                    elif height3 < int(values[f'{model3.names[i3]}_Hn_3']): 
                                        table3.drop(item, axis=0, inplace=True)
                                        area_remove3.append(item)
                                    elif height3 > int(values[f'{model3.names[i3]}_Hx_3']): 
                                        table3.drop(item, axis=0, inplace=True)
                                        area_remove3.append(item)
                                    elif conf3 < int(values[f'{model3.names[i3]}_Conf_3']):
                                        table3.drop(item, axis=0, inplace=True)
                                        area_remove3.append(item)    

                        if values[f'{model3.names[i3]}_3'] == False:
                            if label_name == model3.names[i3]:
                                table3.drop(item, axis=0, inplace=True)
                                area_remove3.append(item)


                    names3 = list(table3['name'])

                    show3 = np.squeeze(result3.render(area_remove3))
                    show3 = cv2.resize(show3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)

                    #ta = time.time()
                    for i3 in range(len(model3.names)):
                        #register_ng = (3002 + i3*2).to_bytes(2, byteorder='big') + b'\x00'

                        if values[f'{model3.names[i3]}_OK_3'] == True:
                            len_name3 = 0
                            for name3 in names3:
                                if name3 == model3.names[i3]:
                                    len_name3 +=1
                            if len_name3 != int(values[f'{model3.names[i3]}_Num_3']):
                                print('NG')
                                #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x03',1)
                                cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam3'].update(value= 'NG', text_color='red')
                                myresult3 = 1
                                

                        elif values[f'{model3.names[i3]}_NG_3'] == True:
                            if model3.names[i3] in names3:
                                print('NG')
                                #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,register_ng,b'\x00\x03',1)
                                cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam3'].update(value= 'NG', text_color='red')    
                                myresult3 = 1         
                                    

                    if myresult3 == 0:
                        print('OK')
                        check_ok = 1
                        #fins_instance.memory_area_write(FinsPLCMemoryAreas().DATA_MEMORY_WORD,(3000).to_bytes(2, byteorder='big') + b'\x00',b'\x00\x03',1)
                        cv2.putText(show3, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
                        window['result_cam3'].update(value= 'OK', text_color='green')

                    imgbytes3 = cv2.imencode('.png',show3)[1].tobytes()
                    window['image3'].update(data= imgbytes3)

                    if check_ok == 1:
                        break



        if event == 'check_model4' and values['check_model4'] == True:
            check_ok=0
            directory4 = 'C:/Check4/'
            if os.listdir(directory4) == []:
                print('folder 4 empty')
            else:
                print('received folder 4')

                for path4 in glob.glob('C:/Check4/*'):
                    name = path4[9:]

                    img4_orgin = cv2.imread(path4)


                    # img4_orgin = cv2.resize(img4_orgin,(640,480))  

                    img4_orgin = cv2.cvtColor(img4_orgin, cv2.COLOR_BGR2RGB)     


                    result4 = model4(img4_orgin,size= 416,conf = values['conf_thres4']/100)

                    table4 = result4.pandas().xyxy[0]

                    area_remove4 = []

                    myresult4 =0 

                    for item in range(len(table4.index)):
                        width4 = table4['xmax'][item] - table4['xmin'][item]
                        height4 = table4['ymax'][item] - table4['ymin'][item]
                        conf4 = table4['confidence'][item] * 100

                        label_name = table4['name'][item]
                        for i4 in range(len(model4.names)):
                            if values[f'{model4.names[i4]}_4'] == True:
    
                                if label_name == model4.names[i4]:
                                    if width4 < int(values[f'{model4.names[i4]}_Wn_4']): 
                                        table4.drop(item, axis=0, inplace=True)
                                        area_remove4.append(item)
                                    elif width4 > int(values[f'{model4.names[i4]}_Wx_4']): 
                                        table4.drop(item, axis=0, inplace=True)
                                        area_remove4.append(item)
                                    elif height4 < int(values[f'{model4.names[i4]}_Hn_4']): 
                                        table4.drop(item, axis=0, inplace=True)
                                        area_remove4.append(item)
                                    elif height4 > int(values[f'{model4.names[i4]}_Hx_4']): 
                                        table4.drop(item, axis=0, inplace=True)
                                        area_remove4.append(item)
                                    elif conf4 < int(values[f'{model4.names[i4]}_Conf_4']):
                                        table4.drop(item, axis=0, inplace=True)
                                        area_remove4.append(item) 
                        if values[f'{model4.names[i4]}_4'] == False:
                            if label_name == model4.names[i4]:
                                table4.drop(item, axis=0, inplace=True)
                                area_remove4.append(item)


                    names4 = list(table4['name'])

                    show4 = np.squeeze(result4.render(area_remove4))
                    show4 = cv2.resize(show4, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)

                    #ta = time.time()
                    for i4 in range(len(model4.names)):

                        if values[f'{model4.names[i4]}_OK_4'] == True:
                            len_name4 = 0
                            for name4 in names4:
                                if name4 == model4.names[i4]:
                                    len_name4 +=1
                            if len_name4 != int(values[f'{model4.names[i4]}_Num_4']):
                                print('NG')
                                cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam4'].update(value= 'NG', text_color='red')
                                myresult4 = 1
                                

                        elif values[f'{model4.names[i4]}_NG_4'] == True:
                            if model4.names[i4] in names4:
                                print('NG')
                                cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,0,255),5)
                                window['result_cam4'].update(value= 'NG', text_color='red')    
                                myresult4 = 1         
                                    

                    if myresult4 == 0:
                        print('OK')
                        check_ok = 1
                        cv2.putText(show4, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 4,(0,255,0),5)
                        window['result_cam4'].update(value= 'OK', text_color='green')

                    imgbytes4 = cv2.imencode('.png',show4)[1].tobytes()
                    window['image4'].update(data= imgbytes4)

                    if check_ok == 1:
                        break

        if event == 'Webcam1':
            #cap1 = cv2.VideoCapture(0)
            recording1 = True


        elif event == 'Stop1':
            recording1 = False 
            imgbytes1 = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes1 = cv2.resize(imgbytes1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes1 = cv2.imencode('.png',imgbytes1)[1].tobytes()
            window['image1'].update(data=imgbytes1)
            window['result_cam1'].update(value='')


        if event == 'Webcam2':
            #cap2 = cv2.VideoCapture(1)
            recording2 = True


        elif event == 'Stop2':
            recording2 = False 
            imgbytes2 = np.zeros([100,100,3],dtype=np.uint8)
            imgbytes2 = cv2.resize(imgbytes2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes2 = cv2.imencode('.png',imgbytes2)[1].tobytes()
            window['image2'].update(data=imgbytes2)
            window['result_cam2'].update(value='')

        if event == 'Webcam3':
            recording3 = True


        elif event == 'Stop3':
            recording3 = False 
            imgbytes3 = np.zeros([300,300,3],dtype=np.uint8)
            imgbytes3 = cv2.resize(imgbytes3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes3 = cv2.imencode('.png',imgbytes3)[3].tobytes()
            window['image3'].update(data=imgbytes3)
            window['result_cam3'].update(value='')


        if event == 'Webcam4':
            recording4 = True


        elif event == 'Stop4':
            recording4 = False 
            imgbytes4 = np.zeros([300,300,3],dtype=np.uint8)
            imgbytes4 = cv2.resize(imgbytes4, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
            imgbytes4 = cv2.imencode('.png',imgbytes4)[3].tobytes()
            window['image4'].update(data=imgbytes4)
            window['result_cam4'].update(value='')


        if recording1:
            # if values['have_model1'] == True:
            #     img1_orgin = my_callback1.image 
            #     img1_orgin = img1_orgin[50:530,70:710]
            #     img1_orgin = img1_orgin.copy()
            #     img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB)                              
            #     result1 = model1(img1_orgin,size= 416,conf= values['conf_thres1']/100)             
            #     table1 = result1.pandas().xyxy[0]
            #     area_remove1 = []

            #     myresult1 =0 

            #     for item in range(len(table1.index)):
            #         width1 = table1['xmax'][item] - table1['xmin'][item]
            #         height1 = table1['ymax'][item] - table1['ymin'][item]
            #         #area1 = width1*height1
            #         label_name = table1['name'][item]
            #         for i1 in range(len(model1.names)):
            #             if values[f'{model1.names[i1]}_1'] == True:
            #                 #if values[f'{model1.names[i1]}_WH'] == True:
            #                 if label_name == model1.names[i1]:
            #                     if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
            #                         table1.drop(item, axis=0, inplace=True)
            #                         area_remove1.append(item)
            #                     elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
            #                         table1.drop(item, axis=0, inplace=True)
            #                         area_remove1.append(item)
            #                     elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
            #                         table1.drop(item, axis=0, inplace=True)
            #                         area_remove1.append(item)
            #                     elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
            #                         table1.drop(item, axis=0, inplace=True)
            #                         area_remove1.append(item)

            #     names1 = list(table1['name'])

            #     show1 = np.squeeze(result1.render(area_remove1))
            #     show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        
            #     #ta = time.time()
            #     for i1 in range(len(model1.names)):
            #         if values[f'{model1.names[i1]}_OK_1'] == True:
            #             len_name1 = 0
            #             for name1 in names1:
            #                 if name1 == model1.names[i1]:
            #                     len_name1 +=1
            #             if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
            #                 print('NG')
            #                 cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam1'].update(value= 'NG', text_color='red')
            #                 myresult1 = 1
            #                 break

            #         if values[f'{model1.names[i1]}_NG_1'] == True:
            #             if model1.names[i1] in names1:
            #                 print('NG')
            #                 cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam1'].update(value= 'NG', text_color='red')    
            #                 myresult1 = 1         
            #                 break    

            #     if myresult1 == 0:
            #         print('OK')
            #         cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            #         window['result_cam1'].update(value= 'OK', text_color='green')
                
            #     imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
            #     window['image1'].update(data= imgbytes1)
            # else:
            img1_orgin = my_callback1.image 
            img1_orgin = img1_orgin[50:530,70:710]
            img1_orgin = img1_orgin.copy()
            img1_orgin = cv2.cvtColor(img1_orgin, cv2.COLOR_BGR2RGB) 
            img1_resize = cv2.resize(img1_orgin,(image_width_display,image_height_display))
            if img1_orgin is not None:
                show1 = img1_resize
                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data=imgbytes1)
                window['result_cam1'].update(value='')


        if recording2:
            # if values['have_model2'] == True:
            #     img2_orgin = my_callback2.image  
            #     img2_orgin = img2_orgin[50:530,70:710]
            #     img2_orgin = img2_orgin.copy()
            #     img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB)                              
            #     result2 = model2(img2_orgin,size= 416,conf= values['conf_thres2']/100)             
            #     table2 = result2.pandas().xyxy[0]
            #     area_remove2 = []

            #     myresult2 =0 

            #     for item in range(len(table2.index)):
            #         width2 = table2['xmax'][item] - table2['xmin'][item]
            #         height2 = table2['ymax'][item] - table2['ymin'][item]
            #         #area2 = width2*height2
            #         label_name = table2['name'][item]
            #         for i2 in range(len(model2.names)):
            #             if values[f'{model2.names[i2]}_2'] == True:
            #                 if label_name == model2.names[i2]:
            #                     if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
            #                         table2.drop(item, axis=0, inplace=True)
            #                         area_remove2.append(item)
            #                     elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
            #                         table2.drop(item, axis=0, inplace=True)
            #                         area_remove2.append(item)
            #                     elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
            #                         table2.drop(item, axis=0, inplace=True)
            #                         area_remove2.append(item)
            #                     elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
            #                         table2.drop(item, axis=0, inplace=True)
            #                         area_remove2.append(item)

            #     names2 = list(table2['name'])

            #     show2 = np.squeeze(result2.render(area_remove2))
            #     show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        
            #     #ta = time.time()
            #     for i2 in range(len(model2.names)):
            #         if values[f'{model2.names[i2]}_OK_2'] == True:
            #             len_name2 = 0
            #             for name2 in names2:
            #                 if name2 == model2.names[i2]:
            #                     len_name2 +=2
            #             if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
            #                 print('NG')
            #                 cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam2'].update(value= 'NG', text_color='red')
            #                 myresult2 = 1
            #                 break

            #         if values[f'{model2.names[i2]}_NG_2'] == True:
            #             if model2.names[i2] in names2:
            #                 print('NG')
            #                 cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam2'].update(value= 'NG', text_color='red')    
            #                 myresult2 = 1         
            #                 break    

            #     if myresult2 == 0:
            #         print('OK')
            #         cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            #         window['result_cam2'].update(value= 'OK', text_color='green')
                
            #     imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
            #     window['image2'].update(data= imgbytes2)
            # else:
            img2_orgin = my_callback2.image  
            img2_orgin = img2_orgin[50:530,70:710]
            img2_orgin = img2_orgin.copy()
            img2_orgin = cv2.cvtColor(img2_orgin, cv2.COLOR_BGR2RGB) 
            img2_resize = cv2.resize(img2_orgin,(image_width_display,image_height_display))
            if img2_orgin is not None:
                show2 = img2_resize
                imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                window['image2'].update(data=imgbytes2)
                window['result_cam2'].update(value='')


        if recording3:
            #if values['have_model3'] == True:
            #     img3_orgin = my_callback3.image 
            #     img3_orgin = img3_orgin[50:530,70:710]
            #     img3_orgin = img3_orgin.copy()
            #     img3_orgin = cv2.cvtColor(img3_orgin, cv2.COLOR_BGR2RGB)                              
            #     result3 = model3(img3_orgin,size= 416,conf= values['conf_thres3']/100)             
            #     table3 = result3.pandas().xyxy[0]
            #     area_remove3 = []

            #     myresult3 =0 

            #     for item in range(len(table3.index)):
            #         width3 = table3['xmax'][item] - table3['xmin'][item]
            #         height3 = table3['ymax'][item] - table3['ymin'][item]

            #         label_name = table3['name'][item]
            #         for i3 in range(len(model3.names)):
            #             if values[f'{model3.names[i3]}_3'] == True:
            #                 if label_name == model3.names[i3]:
            #                     if width3 < int(values[f'{model3.names[i3]}_Wn_3']): 
            #                         table3.drop(item, axis=0, inplace=True)
            #                         area_remove3.append(item)
            #                     elif width3 > int(values[f'{model3.names[i3]}_Wx_3']): 
            #                         table3.drop(item, axis=0, inplace=True)
            #                         area_remove3.append(item)
            #                     elif height3 < int(values[f'{model3.names[i3]}_Hn_3']): 
            #                         table3.drop(item, axis=0, inplace=True)
            #                         area_remove3.append(item)
            #                     elif height3 > int(values[f'{model3.names[i3]}_Hx_3']): 
            #                         table3.drop(item, axis=0, inplace=True)
            #                         area_remove3.append(item)

            #     names3 = list(table3['name'])

            #     show3 = np.squeeze(result3.render(area_remove3))
            #     show3 = cv2.resize(show3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        

            #     for i3 in range(len(model3.names)):
            #         if values[f'{model3.names[i3]}_OK_3'] == True:
            #             len_name3 = 0
            #             for name3 in names3:
            #                 if name3 == model3.names[i3]:
            #                     len_name3 +=1
            #             if len_name3 != int(values[f'{model3.names[i3]}_Num_3']):
            #                 print('NG')
            #                 cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam3'].update(value= 'NG', text_color='red')
            #                 myresult3 = 1
            #                 break

            #         if values[f'{model3.names[i3]}_NG_3'] == True:
            #             if model3.names[i3] in names3:
            #                 print('NG')
            #                 cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam3'].update(value= 'NG', text_color='red')    
            #                 myresult3 = 1         
            #                 break    

            #     if myresult3 == 0:
            #         print('OK')
            #         cv2.putText(show3, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            #         window['result_cam3'].update(value= 'OK', text_color='green')
                
            #     imgbytes3 = cv2.imencode('.png',show3)[1].tobytes()
            #     window['image3'].update(data= imgbytes3)
            # else:
            img3_orgin = my_callback3.image 
            img3_orgin = img3_orgin[50:530,70:710]
            img3_orgin = img3_orgin.copy()
            img3_orgin = cv2.cvtColor(img3_orgin, cv2.COLOR_BGR2RGB) 
            img3_resize = cv2.resize(img3_orgin,(image_width_display,image_height_display))
            if img3_orgin is not None:
                show3 = img3_resize
                imgbytes3 = cv2.imencode('.png',show3)[1].tobytes()
                window['image3'].update(data=imgbytes3)
                window['result_cam3'].update(value='')


        if recording4:
            # if values['have_model4'] == True:
            #     img4_orgin = my_callback4.image  
            #     img4_orgin = img4_orgin[50:530,70:710]
            #     img4_orgin = img4_orgin.copy()
            #     img4_orgin = cv2.cvtColor(img4_orgin, cv2.COLOR_BGR2RGB)                              
            #     result4 = model4(img4_orgin,size= 416,conf= values['conf_thres4']/100)             
            #     table4 = result4.pandas().xyxy[0]
            #     area_remove4 = []

            #     myresult4 =0 

            #     for item in range(len(table4.index)):
            #         width4 = table4['xmax'][item] - table4['xmin'][item]
            #         height4 = table4['ymax'][item] - table4['ymin'][item]
            #         label_name = table4['name'][item]
            #         for i4 in range(len(model4.names)):
            #             if values[f'{model4.names[i4]}_4'] == True:
            #                 if label_name == model4.names[i4]:
            #                     if width4 < int(values[f'{model4.names[i4]}_Wn_4']): 
            #                         table4.drop(item, axis=0, inplace=True)
            #                         area_remove4.append(item)
            #                     elif width4 > int(values[f'{model4.names[i4]}_Wx_4']): 
            #                         table4.drop(item, axis=0, inplace=True)
            #                         area_remove4.append(item)
            #                     elif height4 < int(values[f'{model4.names[i4]}_Hn_4']): 
            #                         table4.drop(item, axis=0, inplace=True)
            #                         area_remove4.append(item)
            #                     elif height4 > int(values[f'{model4.names[i4]}_Hx_4']): 
            #                         table4.drop(item, axis=0, inplace=True)
            #                         area_remove4.append(item)

            #     names4 = list(table4['name'])

            #     show4 = np.squeeze(result4.render(area_remove4))
            #     show4 = cv2.resize(show4, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
        
            #     #ta = time.time()
            #     for i4 in range(len(model4.names)):
            #         if values[f'{model4.names[i4]}_OK_4'] == True:
            #             len_name4 = 0
            #             for name4 in names4:
            #                 if name4 == model4.names[i4]:
            #                     len_name4 +=4
            #             if len_name4 != int(values[f'{model4.names[i4]}_Num_4']):
            #                 print('NG')
            #                 cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam4'].update(value= 'NG', text_color='red')
            #                 myresult4 = 1
            #                 break

            #         if values[f'{model4.names[i4]}_NG_4'] == True:
            #             if model4.names[i4] in names4:
            #                 print('NG')
            #                 cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
            #                 window['result_cam4'].update(value= 'NG', text_color='red')    
            #                 myresult4 = 1         
            #                 break    

            #     if myresult4 == 0:
            #         print('OK')
            #         cv2.putText(show4, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
            #         window['result_cam4'].update(value= 'OK', text_color='green')
                
            #     imgbytes4 = cv2.imencode('.png',show4)[1].tobytes()
            #     window['image4'].update(data= imgbytes4)
            # else:
            img4_orgin = my_callback4.image  
            img4_orgin = img4_orgin[50:530,70:710]
            img4_orgin = img4_orgin.copy()
            img4_orgin = cv2.cvtColor(img4_orgin, cv2.COLOR_BGR2RGB) 
            img4_resize = cv2.resize(img4_orgin,(image_width_display,image_height_display))
            if img4_orgin is not None:
                show4 = img4_resize
                imgbytes4 = cv2.imencode('.png',show4)[1].tobytes()
                window['image4'].update(data=imgbytes4)
                window['result_cam4'].update(value='')




        if event == 'Pic1':
            dir_img1 = sg.popup_get_file('Choose your image 1',file_types=file_name_img,keep_on_top= True)
            if dir_img1 not in ('',None):
                pic1 = Image.open(dir_img1)
                img1_resize = pic1.resize((image_width_display,image_height_display))
                imgbytes1 = ImageTk.PhotoImage(img1_resize)
                window['image1'].update(data= imgbytes1)
                window['Detect1'].update(disabled= False)         

        if event == 'Pic2':
            dir_img2 = sg.popup_get_file('Choose your image 2',file_types=file_name_img,keep_on_top= True)
            if dir_img2 not in ('',None):
                pic2 = Image.open(dir_img2)
                img2_resize = pic2.resize((image_width_display,image_height_display))
                imgbytes2 = ImageTk.PhotoImage(img2_resize)
                window['image2'].update(data=imgbytes2)
                window['Detect2'].update(disabled= False)

        if event == 'Pic3':
            dir_img3 = sg.popup_get_file('Choose your image 3',file_types=file_name_img,keep_on_top= True)
            if dir_img3 not in ('',None):
                pic3 = Image.open(dir_img3)
                img3_resize = pic3.resize((image_width_display,image_height_display))
                imgbytes3 = ImageTk.PhotoImage(img3_resize)
                window['image3'].update(data= imgbytes3)
                window['Detect3'].update(disabled= False)         

        if event == 'Pic4':
            dir_img4 = sg.popup_get_file('Choose your image 4',file_types=file_name_img,keep_on_top= True)
            if dir_img4 not in ('',None):
                pic4 = Image.open(dir_img4)
                img4_resize = pic4.resize((image_width_display,image_height_display))
                imgbytes4 = ImageTk.PhotoImage(img4_resize)
                window['image4'].update(data=imgbytes4)
                window['Detect4'].update(disabled= False)


        if event == 'Change1' or event == 'Change_1':
            list_variable = [[0]*12 for i in range(len(model1.names))]

            for i,item in enumerate(range(len(model1.names))):
                list_variable[i][0] = model1.names[i]

                list_variable[i][1] = values[f'{model1.names[item]}_1']
                list_variable[i][2] = values[f'{model1.names[item]}_OK_1'] 
                list_variable[i][3] = values[f'{model1.names[item]}_Num_1'] 
                list_variable[i][4] = values[f'{model1.names[item]}_NG_1'] 
                list_variable[i][5] = values[f'{model1.names[item]}_Wn_1'] 
                list_variable[i][6] = values[f'{model1.names[item]}_Wx_1'] 
                list_variable[i][7] = values[f'{model1.names[item]}_Hn_1'] 
                list_variable[i][8] = values[f'{model1.names[item]}_Hx_1'] 
                list_variable[i][9] = values[f'{model1.names[item]}_PLC_1'] 
                list_variable[i][10] = values['OK_PLC_1']
                list_variable[i][11] = values[f'{model1.names[item]}_Conf_1'] 

            mypath1 = values['file_weights1']
            model1= torch.hub.load('./levu','custom',path=mypath1,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 
            weight4 = values['file_weights4']
            conf_thres4 = values['conf_thres4'] 

            OK_Cam3 = values['have_save_OK_3']
            OK_Cam4 = values['have_save_OK_4']
            NG_Cam3 = values['have_save_NG_3']
            NG_Cam4 = values['have_save_NG_4']
            Folder_OK_Cam3 = values['save_OK_3']
            Folder_OK_Cam4 = values['save_OK_4']
            Folder_NG_Cam3 = values['save_NG_3']
            Folder_NG_Cam4 = values['save_NG_4']


            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)


            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)
            window['file_weights4'].update(value=weight4)
            window['conf_thres4'].update(value=conf_thres4)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_OK_4'].update(value=OK_Cam4)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['have_save_NG_4'].update(value=NG_Cam4)

            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_OK_4'].update(value=Folder_OK_Cam4)
            window['save_NG_3'].update(value=Folder_NG_Cam3)
            window['save_NG_4'].update(value=Folder_NG_Cam4)

            for i, item in enumerate(range(len(model1.names))):
                for name_label in model1.names:
                    if len(model1.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model1.names[item]}_1'].update(value= list_variable[i][1])
                            window[f'{model1.names[item]}_OK_1'].update(value= list_variable[i][2])
                            window[f'{model1.names[item]}_Num_1'].update(value= list_variable[i][3])
                            window[f'{model1.names[item]}_NG_1'].update(value= list_variable[i][4])
                            window[f'{model1.names[item]}_Wn_1'].update(value= list_variable[i][5])
                            window[f'{model1.names[item]}_Wx_1'].update(value= list_variable[i][6])
                            window[f'{model1.names[item]}_Hn_1'].update(value= list_variable[i][7])
                            window[f'{model1.names[item]}_Hx_1'].update(value= list_variable[i][8])
                            window[f'{model1.names[item]}_PLC_1'].update(value= list_variable[i][9])
                            window['OK_PLC_1'].update(value= list_variable[i][10])
                            window[f'{model1.names[item]}_Conf_1'].update(value= list_variable[i][11])

        if event == 'Change2' or event == 'Change_2':
            list_variable = [[0]*12 for i in range(len(model2.names))]

            for i,item in enumerate(range(len(model2.names))):
                list_variable[i][0] = model2.names[i]

                list_variable[i][1] = values[f'{model2.names[item]}_2']
                list_variable[i][2] = values[f'{model2.names[item]}_OK_2'] 
                list_variable[i][3] = values[f'{model2.names[item]}_Num_2'] 
                list_variable[i][4] = values[f'{model2.names[item]}_NG_2'] 
                list_variable[i][5] = values[f'{model2.names[item]}_Wn_2'] 
                list_variable[i][6] = values[f'{model2.names[item]}_Wx_2'] 
                list_variable[i][7] = values[f'{model2.names[item]}_Hn_2'] 
                list_variable[i][8] = values[f'{model2.names[item]}_Hx_2'] 
                list_variable[i][9] = values[f'{model2.names[item]}_PLC_2'] 
                list_variable[i][10] = values['OK_PLC_2']
                list_variable[i][11] = values[f'{model2.names[item]}_Conf_2'] 

            mypath2 = values['file_weights2']
            model2= torch.hub.load('./levu','custom',path=mypath2,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 
            weight4 = values['file_weights4']
            conf_thres4 = values['conf_thres4'] 

            OK_Cam3 = values['have_save_OK_3']
            OK_Cam4 = values['have_save_OK_4']
            NG_Cam3 = values['have_save_NG_3']
            NG_Cam4 = values['have_save_NG_4']
            Folder_OK_Cam3 = values['save_OK_3']
            Folder_OK_Cam4 = values['save_OK_4']
            Folder_NG_Cam3 = values['save_NG_3']
            Folder_NG_Cam4 = values['save_NG_4']


            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)


            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)
            window['file_weights4'].update(value=weight4)
            window['conf_thres4'].update(value=conf_thres4)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_OK_4'].update(value=OK_Cam4)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['have_save_NG_4'].update(value=NG_Cam4)

            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_OK_4'].update(value=Folder_OK_Cam4)
            window['save_NG_3'].update(value=Folder_NG_Cam3)
            window['save_NG_4'].update(value=Folder_NG_Cam4)

            for i, item in enumerate(range(len(model2.names))):
                for name_label in model2.names:
                    if len(model2.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model2.names[item]}_2'].update(value= list_variable[i][1])
                            window[f'{model2.names[item]}_OK_2'].update(value= list_variable[i][2])
                            window[f'{model2.names[item]}_Num_2'].update(value= list_variable[i][3])
                            window[f'{model2.names[item]}_NG_2'].update(value= list_variable[i][4])
                            window[f'{model2.names[item]}_Wn_2'].update(value= list_variable[i][5])
                            window[f'{model2.names[item]}_Wx_2'].update(value= list_variable[i][6])
                            window[f'{model2.names[item]}_Hn_2'].update(value= list_variable[i][7])
                            window[f'{model2.names[item]}_Hx_2'].update(value= list_variable[i][8])
                            window[f'{model2.names[item]}_PLC_2'].update(value= list_variable[i][9])
                            window['OK_PLC_2'].update(value= list_variable[i][10])
                            window[f'{model2.names[item]}_Conf_2'].update(value= list_variable[i][11])


        if event == 'Change3' or event == 'Change_3':
            list_variable = [[0]*12 for i in range(len(model3.names))]

            for i,item in enumerate(range(len(model3.names))):
                list_variable[i][0] = model3.names[i]

                list_variable[i][1] = values[f'{model3.names[item]}_3']
                list_variable[i][2] = values[f'{model3.names[item]}_OK_3'] 
                list_variable[i][3] = values[f'{model3.names[item]}_Num_3'] 
                list_variable[i][4] = values[f'{model3.names[item]}_NG_3'] 
                list_variable[i][5] = values[f'{model3.names[item]}_Wn_3'] 
                list_variable[i][6] = values[f'{model3.names[item]}_Wx_3'] 
                list_variable[i][7] = values[f'{model3.names[item]}_Hn_3'] 
                list_variable[i][8] = values[f'{model3.names[item]}_Hx_3'] 
                list_variable[i][9] = values[f'{model3.names[item]}_PLC_3'] 
                list_variable[i][10] = values['OK_PLC_3']
                list_variable[i][11] = values[f'{model3.names[item]}_Conf_3'] 
            mypath3 = values['file_weights3']
            model3= torch.hub.load('./levu','custom',path=mypath3,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 
            weight4 = values['file_weights4']
            conf_thres4 = values['conf_thres4'] 

            OK_Cam3 = values['have_save_OK_3']
            OK_Cam4 = values['have_save_OK_4']
            NG_Cam3 = values['have_save_NG_3']
            NG_Cam4 = values['have_save_NG_4']
            Folder_OK_Cam3 = values['save_OK_3']
            Folder_OK_Cam4 = values['save_OK_4']
            Folder_NG_Cam3 = values['save_NG_3']
            Folder_NG_Cam4 = values['save_NG_4']


            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)


            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)
            window['file_weights4'].update(value=weight4)
            window['conf_thres4'].update(value=conf_thres4)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_OK_4'].update(value=OK_Cam4)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['have_save_NG_4'].update(value=NG_Cam4)

            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_OK_4'].update(value=Folder_OK_Cam4)
            window['save_NG_3'].update(value=Folder_NG_Cam3)
            window['save_NG_4'].update(value=Folder_NG_Cam4)

            for i, item in enumerate(range(len(model3.names))):
                for name_label in model3.names:
                    if len(model3.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model3.names[item]}_3'].update(value= list_variable[i][1])
                            window[f'{model3.names[item]}_OK_3'].update(value= list_variable[i][2])
                            window[f'{model3.names[item]}_Num_3'].update(value= list_variable[i][3])
                            window[f'{model3.names[item]}_NG_3'].update(value= list_variable[i][4])
                            window[f'{model3.names[item]}_Wn_3'].update(value= list_variable[i][5])
                            window[f'{model3.names[item]}_Wx_3'].update(value= list_variable[i][6])
                            window[f'{model3.names[item]}_Hn_3'].update(value= list_variable[i][7])
                            window[f'{model3.names[item]}_Hx_3'].update(value= list_variable[i][8])
                            window[f'{model3.names[item]}_PLC_3'].update(value= list_variable[i][9])
                            window['OK_PLC_3'].update(value= list_variable[i][10])
                            window[f'{model3.names[item]}_Conf_3'].update(value= list_variable[i][11])

        if event == 'Change4' or event == 'Change_4':
            list_variable = [[0]*12 for i in range(len(model4.names))]

            for i,item in enumerate(range(len(model4.names))):
                list_variable[i][0] = model4.names[i]

                list_variable[i][1] = values[f'{model4.names[item]}_4']
                list_variable[i][2] = values[f'{model4.names[item]}_OK_4'] 
                list_variable[i][3] = values[f'{model4.names[item]}_Num_4'] 
                list_variable[i][4] = values[f'{model4.names[item]}_NG_4'] 
                list_variable[i][5] = values[f'{model4.names[item]}_Wn_4'] 
                list_variable[i][6] = values[f'{model4.names[item]}_Wx_4'] 
                list_variable[i][7] = values[f'{model4.names[item]}_Hn_4'] 
                list_variable[i][8] = values[f'{model4.names[item]}_Hx_4'] 
                list_variable[i][9] = values[f'{model4.names[item]}_PLC_4'] 
                list_variable[i][10] = values['OK_PLC_4']
                list_variable[i][11] = values[f'{model4.names[item]}_Conf_4'] 

            mypath4 = values['file_weights4']
            model1= torch.hub.load('./levu','custom',path=mypath4,source='local',force_reload=False)
            mychoose = values['choose_model']
            weight1 = values['file_weights1']
            conf_thres1 = values['conf_thres1'] 
            weight2 = values['file_weights2']
            conf_thres2 = values['conf_thres2'] 

            OK_Cam1 = values['have_save_OK_1']
            OK_Cam2 = values['have_save_OK_2']
            NG_Cam1 = values['have_save_NG_1']
            NG_Cam2 = values['have_save_NG_2']
            Folder_OK_Cam1 = values['save_OK_1']
            Folder_OK_Cam2 = values['save_OK_2']
            Folder_NG_Cam1 = values['save_NG_1']
            Folder_NG_Cam2 = values['save_NG_2']

            weight3 = values['file_weights3']
            conf_thres3 = values['conf_thres3'] 
            weight4 = values['file_weights4']
            conf_thres4 = values['conf_thres4'] 

            OK_Cam3 = values['have_save_OK_3']
            OK_Cam4 = values['have_save_OK_4']
            NG_Cam3 = values['have_save_NG_3']
            NG_Cam4 = values['have_save_NG_4']
            Folder_OK_Cam3 = values['save_OK_3']
            Folder_OK_Cam4 = values['save_OK_4']
            Folder_NG_Cam3 = values['save_NG_3']
            Folder_NG_Cam4 = values['save_NG_4']


            window.close() 
            window = make_window(theme)

            window['choose_model'].update(value=mychoose)
            window['file_weights1'].update(value=weight1)
            window['conf_thres1'].update(value=conf_thres1)
            window['file_weights2'].update(value=weight2)
            window['conf_thres2'].update(value=conf_thres2)

            window['have_save_OK_1'].update(value=OK_Cam1)
            window['have_save_OK_2'].update(value=OK_Cam2)
            window['have_save_NG_1'].update(value=NG_Cam1)
            window['have_save_NG_2'].update(value=NG_Cam2)

            window['save_OK_1'].update(value=Folder_OK_Cam1)
            window['save_OK_2'].update(value=Folder_OK_Cam2)
            window['save_NG_1'].update(value=Folder_NG_Cam1)
            window['save_NG_2'].update(value=Folder_NG_Cam2)


            window['choose_model'].update(value=mychoose)
            window['file_weights3'].update(value=weight3)
            window['conf_thres3'].update(value=conf_thres3)
            window['file_weights4'].update(value=weight4)
            window['conf_thres4'].update(value=conf_thres4)

            window['have_save_OK_3'].update(value=OK_Cam3)
            window['have_save_OK_4'].update(value=OK_Cam4)
            window['have_save_NG_3'].update(value=NG_Cam3)
            window['have_save_NG_4'].update(value=NG_Cam4)

            window['save_OK_3'].update(value=Folder_OK_Cam3)
            window['save_OK_4'].update(value=Folder_OK_Cam4)
            window['save_NG_3'].update(value=Folder_NG_Cam3)
            window['save_NG_4'].update(value=Folder_NG_Cam4)

            for i, item in enumerate(range(len(model4.names))):
                for name_label in model4.names:
                    if len(model4.names) <= len(list_variable):
                        if name_label == list_variable[i][0]:

                            window[f'{model4.names[item]}_4'].update(value= list_variable[i][1])
                            window[f'{model4.names[item]}_OK_4'].update(value= list_variable[i][2])
                            window[f'{model4.names[item]}_Num_4'].update(value= list_variable[i][3])
                            window[f'{model4.names[item]}_NG_4'].update(value= list_variable[i][4])
                            window[f'{model4.names[item]}_Wn_4'].update(value= list_variable[i][5])
                            window[f'{model4.names[item]}_Wx_4'].update(value= list_variable[i][6])
                            window[f'{model4.names[item]}_Hn_4'].update(value= list_variable[i][7])
                            window[f'{model4.names[item]}_Hx_4'].update(value= list_variable[i][8])
                            window[f'{model4.names[item]}_PLC_4'].update(value= list_variable[i][9])
                            window['OK_PLC_4'].update(value= list_variable[i][10])
                            window[f'{model4.names[item]}_Conf_4'].update(value= list_variable[i][11])
        if event == 'Detect1':
            print('CAM 1 DETECT')
            t1 = time.time()
            try:
            
                result1 = model1(pic1,size= 416,conf = values['conf_thres1']/100)

                table1 = result1.pandas().xyxy[0]
                print(table1)
                area_remove1 = []

                myresult1 =0 

                for item in range(len(table1.index)):
                    width1 = table1['xmax'][item] - table1['xmin'][item]
                    height1 = table1['ymax'][item] - table1['ymin'][item]
                    conf1 = table1['confidence'][item] * 100

                    #area1 = width1*height1
                    label_name = table1['name'][item]
                    for i1 in range(len(model1.names)):
                        if values[f'{model1.names[i1]}_1'] == True:
                            #if values[f'{model1.names[i1]}_WH'] == True:
                            if label_name == model1.names[i1]:
                                if width1 < int(values[f'{model1.names[i1]}_Wn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif width1 > int(values[f'{model1.names[i1]}_Wx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 < int(values[f'{model1.names[i1]}_Hn_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif height1 > int(values[f'{model1.names[i1]}_Hx_1']): 
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)
                                elif conf1 < int(values[f'{model1.names[i1]}_Conf_1']):
                                    table1.drop(item, axis=0, inplace=True)
                                    area_remove1.append(item)         
                        if values[f'{model1.names[i1]}_1'] == False:
                            if label_name == model1.names[i1]:
                                table1.drop(item, axis=0, inplace=True)
                                area_remove1.append(item)

                names1 = list(table1['name'])

                show1 = np.squeeze(result1.render(area_remove1))
                show1 = cv2.resize(show1, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show1 = cv2.cvtColor(show1, cv2.COLOR_BGR2RGB)
                #ta = time.time()
                for i1 in range(len(model1.names)):
                    if values[f'{model1.names[i1]}_1'] == True:
                        if values[f'{model1.names[i1]}_OK_1'] == True:
                            len_name1 = 0
                            for name1 in names1:
                                if name1 == model1.names[i1]:
                                    len_name1 +=1
                            if len_name1 != int(values[f'{model1.names[i1]}_Num_1']):
                                print('NG')
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')
                                myresult1 = 1
                                break

                        if values[f'{model1.names[i1]}_NG_1'] == True:
                            if model1.names[i1] in names1:
                                print('NG')
                                cv2.putText(show1, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                                window['result_cam1'].update(value= 'NG', text_color='red')    
                                myresult1 = 1         
                                break    

                if myresult1 == 0:
                    print('OK')
                    cv2.putText(show1, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam1'].update(value= 'OK', text_color='green')

                imgbytes1 = cv2.imencode('.png',show1)[1].tobytes()
                window['image1'].update(data= imgbytes1)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',14),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam1 = str(int(t2*1000)) + 'ms'
            window['time_cam1'].update(value= time_cam1, text_color='black') 
            print('---------------------------------------------') 


            
        if event == 'Detect2':
            print('CAM 2 DETECT')
            t1 = time.time()
            try:
                result2 = model2(pic2,size= 416,conf = values['conf_thres2']/100)

                table2 = result2.pandas().xyxy[0]

                area_remove2 = []

                myresult2 =0 

                for item in range(len(table2.index)):
                    width2 = table2['xmax'][item] - table2['xmin'][item]
                    height2 = table2['ymax'][item] - table2['ymin'][item]
                    conf2 = table2['confidence'][item] * 100

                    #area2 = width2*height2
                    label_name = table2['name'][item]
                    for i2 in range(len(model2.names)):
                        if values[f'{model2.names[i2]}_2'] == True:
                            #if values[f'{model2.names[i2]}_WH'] == True:
                            if label_name == model2.names[i2]:
                                if width2 < int(values[f'{model2.names[i2]}_Wn_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif width2 > int(values[f'{model2.names[i2]}_Wx_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif height2 < int(values[f'{model2.names[i2]}_Hn_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif height2 > int(values[f'{model2.names[i2]}_Hx_2']): 
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)
                                elif conf2 < int(values[f'{model2.names[i2]}_Conf_2']):
                                    table2.drop(item, axis=0, inplace=True)
                                    area_remove2.append(item)    

                        if values[f'{model2.names[i2]}_2'] == False:
                            if label_name == model2.names[i2]:
                                table2.drop(item, axis=0, inplace=True)
                                area_remove2.append(item)

                names2 = list(table2['name'])

                show2 = np.squeeze(result2.render(area_remove2))
                show2 = cv2.resize(show2, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show2 = cv2.cvtColor(show2, cv2.COLOR_BGR2RGB)
                #ta = time.time()
                for i2 in range(len(model2.names)):
                    if values[f'{model2.names[i2]}_OK_2'] == True:
                        len_name2 = 0
                        for name2 in names2:
                            if name2 == model2.names[i2]:
                                len_name2 +=1
                        if len_name2 != int(values[f'{model2.names[i2]}_Num_2']):
                            print('NG')
                            cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')
                            myresult2 = 1
                            break

                    if values[f'{model2.names[i2]}_NG_2'] == True:
                        if model2.names[i2] in names2:
                            print('NG')
                            cv2.putText(show2, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam2'].update(value= 'NG', text_color='red')    
                            myresult2 = 1      
                            break    

                if myresult2 == 0:
                    print('OK')
                    cv2.putText(show2, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam2'].update(value= 'OK', text_color='green')

                imgbytes2 = cv2.imencode('.png',show2)[1].tobytes()
                window['image2'].update(data= imgbytes2)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',24),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam2 = str(int(t2*1000)) + 'ms'
            window['time_cam2'].update(value= time_cam2, text_color='black') 
            print('---------------------------------------------') 


        if event == 'Detect3':
            print('CAM 3 DETECT')
            t1 = time.time()
            try:
                result3 = model3(pic3,size= 416,conf = values['conf_thres3']/100)

                table3 = result3.pandas().xyxy[0]

                area_remove3 = []

                myresult3 =0 

                for item in range(len(table3.index)):
                    width3 = table3['xmax'][item] - table3['xmin'][item]
                    height3 = table3['ymax'][item] - table3['ymin'][item]
                    conf3 = table3['confidence'][item] * 100

                    #area3 = width3*height3
                    label_name = table3['name'][item]
                    for i3 in range(len(model3.names)):
                        if values[f'{model3.names[i3]}_3'] == True:
                            #if values[f'{model3.names[i3]}_WH'] == True:
                            if label_name == model3.names[i3]:
                                if width3 < int(values[f'{model3.names[i3]}_Wn_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif width3 > int(values[f'{model3.names[i3]}_Wx_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif height3 < int(values[f'{model3.names[i3]}_Hn_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif height3 > int(values[f'{model3.names[i3]}_Hx_3']): 
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)
                                elif conf3 < int(values[f'{model3.names[i3]}_Conf_3']):
                                    table3.drop(item, axis=0, inplace=True)
                                    area_remove3.append(item)                                      
                        if values[f'{model3.names[i3]}_3'] == False:
                            if label_name == model3.names[i3]:
                                table3.drop(item, axis=0, inplace=True)
                                area_remove3.append(item)

                names3 = list(table3['name'])

                show3 = np.squeeze(result3.render(area_remove3))
                show3 = cv2.resize(show3, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show3 = cv2.cvtColor(show3, cv2.COLOR_BGR2RGB)
                #ta = time.time()
                for i3 in range(len(model3.names)):
                    if values[f'{model3.names[i3]}_OK_3'] == True:
                        len_name3 = 0
                        for name3 in names3:
                            if name3 == model3.names[i3]:
                                len_name3 +=1
                        if len_name3 != int(values[f'{model3.names[i3]}_Num_3']):
                            print('NG')
                            cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam3'].update(value= 'NG', text_color='red')
                            myresult3 = 1
                            break

                    if values[f'{model3.names[i3]}_NG_3'] == True:
                        if model3.names[i3] in names3:
                            print('NG')
                            cv2.putText(show3, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam3'].update(value= 'NG', text_color='red')    
                            myresult3 = 1         
                            break    

                if myresult3 == 0:
                    print('OK')
                    cv2.putText(show3, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam3'].update(value= 'OK', text_color='green')

                imgbytes3 = cv2.imencode('.png',show3)[1].tobytes()
                window['image3'].update(data= imgbytes3)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',34),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam3 = str(int(t2*1000)) + 'ms'
            window['time_cam3'].update(value= time_cam3, text_color='black') 
            print('---------------------------------------------') 


            
        if event == 'Detect4':
            print('CAM 4 DETECT')
            t1 = time.time()
            try:
                result4 = model4(pic4,size= 416,conf = values['conf_thres4']/100)

                table4 = result4.pandas().xyxy[0]

                area_remove4 = []

                myresult4 =0 

                for item in range(len(table4.index)):
                    width4 = table4['xmax'][item] - table4['xmin'][item]
                    height4 = table4['ymax'][item] - table4['ymin'][item]
                    conf4 = table4['confidence'][item] * 100

                    #area4 = width4*height4
                    label_name = table4['name'][item]
                    for i4 in range(len(model4.names)):
                        if values[f'{model4.names[i4]}_4'] == True:
                            #if values[f'{model4.names[i4]}_WH'] == True:
                            if label_name == model4.names[i4]:
                                if width4 < int(values[f'{model4.names[i4]}_Wn_4']): 
                                    table4.drop(item, axis=0, inplace=True)
                                    area_remove4.append(item)
                                elif width4 > int(values[f'{model4.names[i4]}_Wx_4']): 
                                    table4.drop(item, axis=0, inplace=True)
                                    area_remove4.append(item)
                                elif height4 < int(values[f'{model4.names[i4]}_Hn_4']): 
                                    table4.drop(item, axis=0, inplace=True)
                                    area_remove4.append(item)
                                elif height4 > int(values[f'{model4.names[i4]}_Hx_4']): 
                                    table4.drop(item, axis=0, inplace=True)
                                    area_remove4.append(item)
                                elif conf4 < int(values[f'{model4.names[i4]}_Conf_4']):
                                    table4.drop(item, axis=0, inplace=True)
                                    area_remove4.append(item)                                     
                        if values[f'{model4.names[i4]}_4'] == False:
                            if label_name == model4.names[i4]:
                                table4.drop(item, axis=0, inplace=True)
                                area_remove4.append(item)

                names4 = list(table4['name'])

                show4 = np.squeeze(result4.render(area_remove4))
                show4 = cv2.resize(show4, (image_width_display,image_height_display), interpolation = cv2.INTER_AREA)
                show4 = cv2.cvtColor(show4, cv2.COLOR_BGR2RGB)
                #ta = time.time()
                for i4 in range(len(model4.names)):
                    if values[f'{model4.names[i4]}_OK_4'] == True:
                        len_name4 = 0
                        for name4 in names4:
                            if name4 == model4.names[i4]:
                                len_name4 +=1
                        if len_name4 != int(values[f'{model4.names[i4]}_Num_4']):
                            print('NG')
                            cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam4'].update(value= 'NG', text_color='red')
                            myresult4 = 1
                            break

                    if values[f'{model4.names[i4]}_NG_4'] == True:
                        if model4.names[i4] in names4:
                            print('NG')
                            cv2.putText(show4, 'NG',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,0,255),5)
                            window['result_cam4'].update(value= 'NG', text_color='red')    
                            myresult4 = 1      
                            break    

                if myresult4 == 0:
                    print('OK')
                    cv2.putText(show4, 'OK',(result_width_display,result_height_display),cv2.FONT_HERSHEY_COMPLEX, 3,(0,255,0),5)
                    window['result_cam4'].update(value= 'OK', text_color='green')

                imgbytes4 = cv2.imencode('.png',show4)[1].tobytes()
                window['image4'].update(data= imgbytes4)

            
            except:
                print(traceback.format_exc())
                sg.popup_annoying("Don't have image or parameter wrong", font=('Helvetica',44),text_color='red')
            
            t2 = time.time() - t1
            print(t2)
            time_cam4 = str(int(t2*1000)) + 'ms'
            window['time_cam4'].update(value= time_cam4, text_color='black') 
            print('---------------------------------------------') 

    window.close() 

except Exception as e:
    print(traceback.print_exc())
    str_error = str(e)
    sg.popup(str_error,font=('Helvetica',15), text_color='red',keep_on_top= True)
            